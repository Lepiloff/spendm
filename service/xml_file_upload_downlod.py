import re
from openpyxl import load_workbook

from apps.vendors.models import RfiParticipation, Vendors, Rfis, ParentCategories, RfiParticipationStatus, \
    CompanyGeneralInfoQuestion

from rest_framework.response import Response


header_cols_n_scoring_round = {
                               '1': ["E", "F", "G", "P", "Q", "R", "S", "T"],
                               '2': ["E", "F", "G", "U", "V", "W", "X", "Y"],
                               }


def header_cols_for_scoring_round(scoring_round):
    result = header_cols_n_scoring_round.get(scoring_round)
    return result


pc_to_modules_assign = {
                            "Strategic Sourcing": ['COMMON S2P ', 'COMMON SOURCING – SXM ', 'SERVICES ', 'SOURCING '],
                            "Supplier Management": ['COMMON S2P ', 'COMMON SOURCING – SXM ', 'SERVICES ', 'SXM '],
                            "Spend Analytics": ['COMMON S2P ', 'SERVICES ', 'Spend Analytics '],
                            "Contract Management": ['COMMON S2P ', 'SERVICES ', 'CLM '],
                            "e-Procurement": ['COMMON S2P ', 'SERVICES ', 'eProcurement '],
                            "Invoice-to-Pay": ['COMMON S2P ', 'SERVICES ', 'I2P '],
                            "AP Automation": ['COMMON S2P ', 'SERVICES ', 'I2P ', 'AP '],
                            "Strategic Procurement Technologies": ['COMMON S2P ', 'COMMON SOURCING – SXM ', 'SERVICES ',
                                                                   'SOURCING ', 'SXM ', 'Spend Analytics ', 'CLM '],
                            "Procure-to-Pay": ['COMMON S2P ', 'SERVICES ', 'eProcurement ', 'I2P '],
                            "Source-to-Pay": ['COMMON S2P ', 'COMMON SOURCING – SXM ', 'SERVICES ',
                                              'SOURCING ', 'SXM ', 'Spend Analytics ', 'CLM ', 'eProcurement ', 'I2P ']
                             }


class InvalidFormatException(Exception):
    # do not remove !
    pass


class InvalidRoundException(Exception):
    # do not remove !
    pass


class InvalidCharactersException(Exception):
    # do not remove !
    pass


class CIEmptyException(Exception):
    # do not remove !
    pass


def company_info_parser(file, analyst, _round, vendor):
    """
    Get Company information response
    :param file:
    :return:
    """
    workbook = load_workbook(filename=file)
    sheet = workbook["Company Information"]
    company_general_info = []
    for row in sheet.iter_rows(min_row=5, max_row=34, values_only=False):
        company_info = {}
        for cell in row:
            if cell.column_letter == "B":
                company_info['question'] = cell.value
            if cell.column_letter == "C":
                company_info['answer'] = cell.value
        company_info.update(company_info)
        company_general_info.append(company_info)

    if not analyst:  # Check that file send by vendor
        # Check that CI answer stored in DB yet (at list one)
        ci_db = check_ci_exist_in_db(_round, vendor)
        # Check company info from excel file
        ci_file = get_ci_from_excel_file(company_general_info)
        if not ci_db and not ci_file:
            raise CIEmptyException("The company information can't be empty.")
    return company_general_info


def check_ci_exist_in_db(_round, vendor):
    """Check that CI answer stored in DB yet (at list one)"""
    exist_company_question = CompanyGeneralInfoQuestion.objects.filter(rfi=_round)
    ci_exist = False

    if exist_company_question:
        for q in exist_company_question:
            if q.answer_to_question.filter(vendor=vendor):  # check if answer is exist and not None
                _a = (q.answer_to_question.filter(vendor=vendor).first())
                if _a.answer:
                    ci_exist = True
    return ci_exist

def get_ci_from_excel_file(company_general_info):
    """"Check company info from excel file"""
    ci_exist = False
    for i in company_general_info:
        if i.get('answer'):
            ci_exist = True
    return ci_exist


def get_excel_file_current_pc_for_parsing(pml=None):
    modules_with_pc = (dict((option, pc_to_modules_assign[option]) for option in pml if option in pc_to_modules_assign))
    pc = [value for key, value in modules_with_pc.items()]
    # Get unique pc depending on the active modules in the round for a particular vendor
    unique_pc = set(x for element in pc for x in element)
    return unique_pc


paren_category_ = [
                            "COMMON S2P ",
                            "COMMON SOURCING – SXM ",
                            "SERVICES ",
                            "SOURCING ",
                            "SXM ",
                            "Spend Analytics ",
                            "CLM ",
                            "eProcurement ",
                            "I2P "
                             ]

def get_pc_coordinate(file):
    pc_coordinate = {}
    workbook = load_workbook(filename=file)
    sheet = workbook["RFI"]
    for row_cells in sheet.iter_rows(min_row=4):
        for cell in row_cells:
            if cell.value in paren_category_:
                pc_coordinate[cell.value] = cell.coordinate
    return pc_coordinate


def get_full_excel_file_response(file, context):
    """ return full response from excel file (exclude PC not participate to vendor in current round)"""
    # Get data from url context
    scoring_round = int(context.get('f_scoring_round'))
    analyst = context.get('analyst')
    rfiid = context.get('rfiid')
    vendor_id = context.get('vendor')
    vendor = Vendors.objects.get(vendorid=vendor_id)
    _round = Rfis.objects.get(rfiid=rfiid)

    # Check that round name in file name match with active round
    file_name_round = Rfis.objects.get(rfiid=context.get('f_round'))
    if not file_name_round.active:
        raise InvalidRoundException("Round from file name is not active")

    # check vendor have participated module in current round
    participation = RfiParticipation.objects.filter(rfi=_round, vendor=vendor, active=True)
    if not participation:
        raise InvalidRoundException("Vendor have't active participated module in current round")

    participate_module = RfiParticipation.objects.filter(vendor=vendor, rfi=_round, active=True)  # Get vendor active module
    participate_module_list = [element.m.module_name for element in participate_module]
    unique_pc = get_excel_file_current_pc_for_parsing(pml=participate_module_list)  # Get unique PC for future processing
    response = []
    # Get Company information response
    company_info = company_info_parser(file, analyst, _round, vendor)
    response.append({"Company_info": company_info})
    # Calculate pc status
    pc_status = []
    workbook = load_workbook(filename=file)
    pc_participate_list = []
    status_info = {}
    # Get coordinate of PC
    pc_coordinate = get_pc_coordinate(file)
    for element in unique_pc:

        pc_participate_list.append({pc_to_function_name.get(element): pc_coordinate.get(element)})
    for pc in pc_participate_list:
        if pc:
            if None in pc: # check that dict not contain None values as a key
                continue
            else:
                # dict.keys return a set, using slicing list to get "naked" function (key) and coordinate (value)
                function_to_call = (list(pc.keys())[0])
                coordinate = (list(pc.values())[0])
                data = function_to_call(workbook, coordinate, scoring_round)
                response.append(data)
                # check PC contain not null value (curent status)
                pc_st = current_score_data(data, vendor, _round, scoring_round, analyst)
                pc_status.append(pc_st)
                # get info on who filled in the information in the element
                _status_info = from_vendor_analyst(data)
                status_info.update(_status_info)
    response.append({'Status_info': status_info})
    # s - current round
    s = {"status": pc_status, 'scoring_round': scoring_round}
    s.update({"Company info": True})
    #  old_pc_st - previous round (s-1)
    old_pc_st = {}
    if scoring_round != 1:
        old_pc_st = past_score_not_all_element_is_null(vendor, _round, scoring_round, unique_pc)
        old_pc_st.update({"Company info": True})
    # Check if last round isn't exist in db skip old_pc_st in response
    if len(old_pc_st) == 0:
        response.append({'Scoring_round_info': [s]})
    else:
        response.append({'Scoring_round_info': [s, old_pc_st]})
    # add to response current scoring round from file name
    response.append({"Scoring_round_current": scoring_round})
    return response


def from_vendor_analyst(data):
    # get info on who filled in the information in the element
    pc = data.get('Parent Category')
    category_data = data.get('Category')
    _status_info = {"vendor": False, "analyst": False}
    for data in category_data:
        for category, values in data.items():  # Get category name
            for subcats in values:
                for subcat, element_list in subcats.items():  # Get subcategory name
                    for element in element_list:  # Get element info
                        self_score = element.get('Self-Score')
                        self_description = element.get('Self-Description')
                        sm_score = element.get('SM score')
                        analyst_notes = element.get('Analyst notes')
                        from_vendor = (self_score, self_description)
                        from_analytic = (sm_score, analyst_notes)

                        if all(from_vendor):
                            _status_info['vendor'] = True
                        if all(from_analytic):
                            _status_info['analyst'] = True
    status_info = {pc: _status_info}
    return status_info


def current_score_data(data, vendor, _round, scoring_round, analyst):

    """
    For curent round
    Check that at list one element pair (self_score/self_description; sm_score/analyst_notes) are not empty.
    That means we can set rfi_part_status to PC as positive digit(1 for first scoring round etc.)
    :param data:
    :return:
    """
    pc = data.get('Parent Category')
    category_data = data.get('Category')
    pc_status = {}
    for data in category_data:
        for category, values in data.items():  # Get category name
            for subcats in values:
                for subcat, element_list in subcats.items():  # Get subcategory name
                    for element in element_list:  # Get element info
                        self_score = element.get('Self-Score')
                        self_description = element.get('Self-Description')
                        sm_score = element.get('SM score')
                        analyst_notes = element.get('Analyst notes')
                        from_vendor = (self_score, self_description)
                        from_analytic = (sm_score, analyst_notes)

                        # If current scoring round exist in db and data consist new information return False
                        # that mean possibility to update old data in DB then change switcher status
                        # s_r_e = scoring_round_exist(pc, vendor, _round)
                        # if s_r_e:
                        #     if all(from_vendor) or all(from_analytic):
                        #         pc_status[pc] = True
                        #         return pc_status

                        # For first scoring round
                        if scoring_round == 1 and not analyst:
                            if all(from_vendor):
                                pc_status[pc] = False
                                return pc_status
                            else:
                                pc_status[pc] = "No data"
                                return pc_status

                        if scoring_round == 1 and analyst:
                            if all(from_analytic) and not all(from_vendor):
                                pc_status[pc] = "-"
                                return pc_status
                            if not all(from_analytic):
                                pc_status[pc] = 'No data'
                                return pc_status

                        s_r_e = scoring_round_exist(pc, vendor, _round)
                        if s_r_e:
                            pc_status[pc] = True
                            return pc_status

                        # # If mailer is vendor and element not empty
                        # if not analyst and all(from_vendor):
                        #     pc_status[pc] = False
                        #     return pc_status
                        # # If mailer is analyst and element not empty
                        # elif analyst and all(from_analytic):
                        #     pc_status[pc] = False
                        #     return pc_status
                        #
                        # elif all(from_vendor) and all(from_analytic):
                        #     pc_status[pc] = True
                        #     return pc_status
                        #
                        # elif (all(from_vendor) and not all(from_analytic)) \
                        #         or (all(from_analytic) and not all(from_vendor)):
                        #     pc_status[pc] = "-"
                        #     return pc_status
                        # else:
                        #     pc_status[pc] = "No data"
                        #     return pc_status

    return pc_status


def scoring_round_exist(pc, vendor, _round):
    # Check that scoring round yet in DB
    pc_obj = ParentCategories.objects.get(parent_category_name=pc)
    status = (RfiParticipationStatus.objects.filter(vendor=vendor, rfi=_round, pc=pc_obj).values(
        "last_vendor_response", "last_analyst_response"))
    for s in status:
        if s.get('last_vendor_response') != 0:
            if s.get('last_analyst_response') != 0:
                return True
        else:
            return False


def past_score_not_all_element_is_null(vendor, round, scoring_round, unique_pc):
    """
    For previous scoring round
    Check that at list one element pair (self_score/self_description; sm_score/analyst_notes) are not empty.
    :param data:
    :return:
    """
    # TODO check difference between previous round and current data
    scoring_round_info = {}
    previews_scoring_status = []
    for pc in unique_pc:
        s_r_n = scoring_round - 1
        pc_obj = ParentCategories.objects.get(parent_category_name=pc)
        status = (RfiParticipationStatus.objects.filter(vendor=vendor, rfi=round, pc=pc_obj).values("last_vendor_response", "last_analyst_response"))
        for s in status:
            if s.get('last_vendor_response') != 0:
                if s.get('last_analyst_response') != 0:
                    previews_scoring_status.append({pc: True})
            else:
                previews_scoring_status.append({pc: False})
    scoring_round_info.update({'status': previews_scoring_status})
    scoring_round_info.update({'scoring_round': s_r_n})
    return scoring_round_info


def subcategory_element_response_create(scoring_round, min_row, max_row, sheet=None, to_category_info=None, sub_category=None):
    to_sub_category_info = []
    header_cols = header_cols_for_scoring_round(str(scoring_round))
    for row in sheet.iter_rows(min_row=min_row, max_row=max_row, values_only=False):
        element_info = {}
        for cell in row:
            if cell.column_letter in header_cols:
                # if if is used to change the name of the element in the not first scoring round,
                # as the file columns for the second round have a different name as a first round
                epn = sheet[f'{cell.column_letter}2'].value  # 2 - because second row with  row name
                if epn == "Self-Score (2)":
                    epn = "Self-Score"
                elif epn == "SM score (2)":
                    epn = "SM score"
                elif epn == "Reasoning":
                    epn = "Self-Description"
                elif epn == "Analyst notes (2)":
                    epn = "Analyst notes"
                element_info[epn] = cell.value
        to_sub_category_info.append(element_info)
        non_latin_characters(to_sub_category_info)
    to_category_info.append({sub_category: to_sub_category_info})


def non_latin_characters(list_):
    key_fields = ['Self-Description', 'Analyst notes']
    for dict_ in list_:
        for key in dict_:
            if key in key_fields:
                if dict_[key] is not None:
                    if not re.match(r'^[a-zA-Z0-9,.!? -/*()]*$', dict_[key]):
                        raise InvalidCharactersException("The system detected that the data is not in English. "
                                                         "Please correct the error and try again.")


def get_split_coordinate(coordinate, bias):
    column_letter = coordinate[0]
    row_num = int(coordinate[1:]) + bias
    return (column_letter, row_num)


def common_s2p_category_response_create(workbook, coordinate, scoring_round):
    """
    Create response from COMMON_S2P parsing
    :param file:
    :return:
    """
    sheet = workbook["RFI"]
    pc_response = {}  # crete response with all element information for each subcat for each category in PC
    pc_response.update({"Parent Category": sheet[coordinate].value})
    category_list = []  # list of all category in PC with subcat info and element data

    # ANALYTICS CATEGORY
    cl, rn = get_split_coordinate(coordinate, 1)
    category_name = sheet[f"{cl}{rn}"].value
    category_coordinate = sheet[f"{cl}{rn}"].coordinate
    info_to_subcat_to_cat = {}
    to_category_info = []  # list of all subcat wit element info

    # Subcategory Data Schema
    cl, rn = get_split_coordinate(category_coordinate, 1)
    sub_category = sheet[f"{cl}{rn}"].value
    sub_category_coordinate = sheet[f"{cl}{rn}"].coordinate

    _, min_row = get_split_coordinate(sub_category_coordinate, 1)
    subcategory_element_response_create(scoring_round, min_row=min_row, max_row=(min_row + 5), sheet=sheet, to_category_info=to_category_info,
                                        sub_category=sub_category)
    # Subcategory Data Management
    cl, rn = get_split_coordinate(category_coordinate, 10)
    sub_category = sheet[f"{cl}{rn}"].value
    sub_category_coordinate = sheet[f"{cl}{rn}"].coordinate

    _, min_row = get_split_coordinate(sub_category_coordinate, 1)
    subcategory_element_response_create(scoring_round, min_row=min_row, max_row=(min_row + 4), sheet=sheet, to_category_info=to_category_info,
                                        sub_category=sub_category)
    # Subcategory Metric Management
    cl, rn = get_split_coordinate(category_coordinate, 18)
    sub_category = sheet[f"{cl}{rn}"].value
    sub_category_coordinate = sheet[f"{cl}{rn}"].coordinate

    _, min_row = get_split_coordinate(sub_category_coordinate, 1)
    subcategory_element_response_create(scoring_round, min_row=min_row, max_row=(min_row + 4), sheet=sheet, to_category_info=to_category_info,
                                        sub_category=sub_category)

    # Subcategory Reporting
    cl, rn = get_split_coordinate(category_coordinate, 26)
    sub_category = sheet[f"{cl}{rn}"].value
    sub_category_coordinate = sheet[f"{cl}{rn}"].coordinate
    _, min_row = get_split_coordinate(sub_category_coordinate, 1)
    subcategory_element_response_create(scoring_round, min_row=min_row, max_row=(min_row + 10), sheet=sheet, to_category_info=to_category_info,
                                        sub_category=sub_category)

    info_to_subcat_to_cat.update({category_name: to_category_info})  # Aggregate info for each Subcat at CAREGORY
    category_list.append(info_to_subcat_to_cat)

    # CONFIGURABILITY CATEGORY
    cl, rn = get_split_coordinate(coordinate, 41)
    category_name = sheet[f"{cl}{rn}"].value
    category_coordinate = sheet[f"{cl}{rn}"].coordinate
    info_to_subcat_to_cat = {}
    to_category_info = []  # list of all subcat wit element info

    # Subcategory Globalization
    cl, rn = get_split_coordinate(category_coordinate, 1)
    sub_category = sheet[f"{cl}{rn}"].value
    sub_category_coordinate = sheet[f"{cl}{rn}"].coordinate
    _, min_row = get_split_coordinate(sub_category_coordinate, 1)
    subcategory_element_response_create(scoring_round, min_row=min_row, max_row=(min_row + 6), sheet=sheet, to_category_info=to_category_info,
                                        sub_category=sub_category)

    # Subcategory Organizational Modeling
    cl, rn = get_split_coordinate(category_coordinate, 11)
    sub_category = sheet[f"{cl}{rn}"].value
    sub_category_coordinate = sheet[f"{cl}{rn}"].coordinate
    _, min_row = get_split_coordinate(sub_category_coordinate, 1)
    subcategory_element_response_create(scoring_round, min_row=min_row, max_row=(min_row + 5), sheet=sheet, to_category_info=to_category_info,
                                        sub_category=sub_category)

    # Subcategory Personalization
    cl, rn = get_split_coordinate(category_coordinate, 20)
    sub_category = sheet[f"{cl}{rn}"].value
    sub_category_coordinate = sheet[f"{cl}{rn}"].coordinate
    _, min_row = get_split_coordinate(sub_category_coordinate, 1)
    subcategory_element_response_create(scoring_round, min_row=min_row, max_row=(min_row + 4), sheet=sheet, to_category_info=to_category_info,
                                        sub_category=sub_category)

    # Subcategory Project Management
    cl, rn = get_split_coordinate(category_coordinate, 28)
    sub_category = sheet[f"{cl}{rn}"].value
    sub_category_coordinate = sheet[f"{cl}{rn}"].coordinate
    _, min_row = get_split_coordinate(sub_category_coordinate, 1)
    subcategory_element_response_create(scoring_round, min_row=min_row, max_row=(min_row + 4), sheet=sheet, to_category_info=to_category_info,
                                        sub_category=sub_category)

    # Subcategory Workflow
    cl, rn = get_split_coordinate(category_coordinate, 36)
    sub_category = sheet[f"{cl}{rn}"].value
    sub_category_coordinate = sheet[f"{cl}{rn}"].coordinate
    _, min_row = get_split_coordinate(sub_category_coordinate, 1)
    subcategory_element_response_create(scoring_round, min_row=min_row, max_row=(min_row + 6), sheet=sheet, to_category_info=to_category_info,
                                        sub_category=sub_category)

    info_to_subcat_to_cat.update({category_name: to_category_info})  # Aggregate info for each Subcat at CAREGORY
    category_list.append(info_to_subcat_to_cat)

    # SUPPLIER PORTAL CATEGORY
    cl, rn = get_split_coordinate(coordinate, 87)
    category_name = sheet[f"{cl}{rn}"].value
    category_coordinate = sheet[f"{cl}{rn}"].coordinate
    info_to_subcat_to_cat = {}
    to_category_info = []  # list of all subcat wit element info


    # Subcategory Account Management
    cl, rn = get_split_coordinate(category_coordinate, 1)
    sub_category = sheet[f"{cl}{rn}"].value
    sub_category_coordinate = sheet[f"{cl}{rn}"].coordinate
    _, min_row = get_split_coordinate(sub_category_coordinate, 1)
    subcategory_element_response_create(scoring_round, min_row=min_row, max_row=(min_row + 4), sheet=sheet, to_category_info=to_category_info,
                                        sub_category=sub_category)

    # Subcategory Document Management
    cl, rn = get_split_coordinate(category_coordinate, 9)
    sub_category = sheet[f"{cl}{rn}"].value
    sub_category_coordinate = sheet[f"{cl}{rn}"].coordinate
    _, min_row = get_split_coordinate(sub_category_coordinate, 1)
    subcategory_element_response_create(scoring_round, min_row=min_row, max_row=(min_row + 1), sheet=sheet, to_category_info=to_category_info,
                                        sub_category=sub_category)

    # Subcategory Information Management
    cl, rn = get_split_coordinate(category_coordinate, 14)
    sub_category = sheet[f"{cl}{rn}"].value
    sub_category_coordinate = sheet[f"{cl}{rn}"].coordinate
    _, min_row = get_split_coordinate(sub_category_coordinate, 1)
    subcategory_element_response_create(scoring_round, min_row=min_row, max_row=(min_row + 2), sheet=sheet, to_category_info=to_category_info,
                                        sub_category=sub_category)

    info_to_subcat_to_cat.update({category_name: to_category_info})  # Aggregate info for each Subcat at CAREGORY
    category_list.append(info_to_subcat_to_cat)

    # SXM CATEGORY
    cl, rn = get_split_coordinate(coordinate, 107)
    category_name = sheet[f"{cl}{rn}"].value
    category_coordinate = sheet[f"{cl}{rn}"].coordinate
    info_to_subcat_to_cat = {}
    to_category_info = []  # list of all subcat wit element info

    # Subcategory Supplier Information Management
    cl, rn = get_split_coordinate(category_coordinate, 1)
    sub_category = sheet[f"{cl}{rn}"].value
    sub_category_coordinate = sheet[f"{cl}{rn}"].coordinate
    _, min_row = get_split_coordinate(sub_category_coordinate, 1)
    subcategory_element_response_create(scoring_round, min_row=min_row, max_row=(min_row + 6), sheet=sheet, to_category_info=to_category_info,
                                        sub_category=sub_category)
    # # Subcategory Supply Intelligence
    cl, rn = get_split_coordinate(category_coordinate, 11)
    sub_category = sheet[f"{cl}{rn}"].value
    sub_category_coordinate = sheet[f"{cl}{rn}"].coordinate
    _, min_row = get_split_coordinate(sub_category_coordinate, 1)
    subcategory_element_response_create(scoring_round, min_row=min_row, max_row=(min_row + 2), sheet=sheet, to_category_info=to_category_info,
                                        sub_category=sub_category)

    info_to_subcat_to_cat.update({category_name: to_category_info})  # Aggregate info for each Subcat at CAREGORY
    category_list.append(info_to_subcat_to_cat)

    # TECHNOLOGY CATEGORY
    cl, rn = get_split_coordinate(coordinate, 124)
    category_name = sheet[f"{cl}{rn}"].value
    category_coordinate = sheet[f"{cl}{rn}"].coordinate
    info_to_subcat_to_cat = {}
    to_category_info = []  # list of all subcat wit element info

    # Subcategory Automation
    cl, rn = get_split_coordinate(category_coordinate, 1)
    sub_category = sheet[f"{cl}{rn}"].value
    sub_category_coordinate = sheet[f"{cl}{rn}"].coordinate
    _, min_row = get_split_coordinate(sub_category_coordinate, 1)
    subcategory_element_response_create(scoring_round, min_row=min_row, max_row=(min_row + 6), sheet=sheet, to_category_info=to_category_info,
                                        sub_category=sub_category)
    # Subcategory Core Platform
    cl, rn = get_split_coordinate(category_coordinate, 11)
    sub_category = sheet[f"{cl}{rn}"].value
    sub_category_coordinate = sheet[f"{cl}{rn}"].coordinate
    _, min_row = get_split_coordinate(sub_category_coordinate, 1)
    subcategory_element_response_create(scoring_round, min_row=min_row, max_row=(min_row + 16), sheet=sheet, to_category_info=to_category_info,
                                        sub_category=sub_category)

    # Subcategory Data Management
    cl, rn = get_split_coordinate(category_coordinate, 31)
    sub_category = sheet[f"{cl}{rn}"].value
    sub_category_coordinate = sheet[f"{cl}{rn}"].coordinate
    _, min_row = get_split_coordinate(sub_category_coordinate, 1)
    subcategory_element_response_create(scoring_round, min_row=min_row, max_row=(min_row + 14), sheet=sheet, to_category_info=to_category_info,
                                        sub_category=sub_category)

    # Subcategory Document Management
    cl, rn = get_split_coordinate(category_coordinate, 49)
    sub_category = sheet[f"{cl}{rn}"].value
    sub_category_coordinate = sheet[f"{cl}{rn}"].coordinate
    _, min_row = get_split_coordinate(sub_category_coordinate, 1)
    subcategory_element_response_create(scoring_round, min_row=min_row, max_row=(min_row + 4), sheet=sheet, to_category_info=to_category_info,
                                        sub_category=sub_category)

    # Subcategory Emerging Technology
    cl, rn = get_split_coordinate(category_coordinate, 57)
    sub_category = sheet[f"{cl}{rn}"].value
    sub_category_coordinate = sheet[f"{cl}{rn}"].coordinate
    _, min_row = get_split_coordinate(sub_category_coordinate, 1)
    subcategory_element_response_create(scoring_round, min_row=min_row, max_row=(min_row + 7), sheet=sheet, to_category_info=to_category_info,
                                        sub_category=sub_category)

    # Subcategory Standards and Integrations
    cl, rn = get_split_coordinate(category_coordinate, 68)
    sub_category = sheet[f"{cl}{rn}"].value
    sub_category_coordinate = sheet[f"{cl}{rn}"].coordinate
    _, min_row = get_split_coordinate(sub_category_coordinate, 1)
    subcategory_element_response_create(scoring_round, min_row=min_row, max_row=(min_row + 12), sheet=sheet, to_category_info=to_category_info,
                                        sub_category=sub_category)

    # Subcategory UX Layer
    cl, rn = get_split_coordinate(category_coordinate, 84)
    sub_category = sheet[f"{cl}{rn}"].value
    sub_category_coordinate = sheet[f"{cl}{rn}"].coordinate
    _, min_row = get_split_coordinate(sub_category_coordinate, 1)
    subcategory_element_response_create(scoring_round, min_row=min_row, max_row=(min_row + 6), sheet=sheet, to_category_info=to_category_info,
                                        sub_category=sub_category)
    info_to_subcat_to_cat.update({category_name: to_category_info})  # Aggregate info for each Subcat at CAREGORY
    category_list.append(info_to_subcat_to_cat)

    pc_response.update({"Category": category_list})
    return pc_response


def common_sourcing_sxm_response_create(workbook, coordinate, scoring_round):
    """
    Create response from COMMON SOURCING SXM parsing
    :param file:
    :return:
    """
    sheet = workbook["RFI"]
    pc_response = {}  # crete response with all element information for each subcat for each category in PC
    pc_response.update({"Parent Category": sheet[coordinate].value})
    category_list = []  # list of all category in PC with subcat info and element data

    # CONTINGENT WORKFORCE / SERVICES PROCUREMENT CATEGORY
    cl, rn = get_split_coordinate(coordinate, 1)
    category_name = sheet[f"{cl}{rn}"].value
    category_coordinate = sheet[f"{cl}{rn}"].coordinate
    info_to_subcat_to_cat = {}
    to_category_info = []  # list of all subcat wit element info

    # Subcategory Contingent Workforce Management
    cl, rn = get_split_coordinate(category_coordinate, 1)
    sub_category = sheet[f"{cl}{rn}"].value
    sub_category_coordinate = sheet[f"{cl}{rn}"].coordinate
    _, min_row = get_split_coordinate(sub_category_coordinate, 1)
    subcategory_element_response_create(scoring_round, min_row=min_row, max_row=(min_row + 2), sheet=sheet, to_category_info=to_category_info,
                                        sub_category=sub_category)

    info_to_subcat_to_cat.update({category_name: to_category_info})  # Aggregate info for each Subcat at CAREGORY
    category_list.append(info_to_subcat_to_cat)

    # PERFOMANCE MANAGEMENT CATEGORY

    cl, rn = get_split_coordinate(coordinate, 8)
    category_name = sheet[f"{cl}{rn}"].value
    category_coordinate = sheet[f"{cl}{rn}"].coordinate
    info_to_subcat_to_cat = {}
    to_category_info = []  # list of all subcat wit element info

    # Subcategory Supplier Performance Management
    cl, rn = get_split_coordinate(category_coordinate, 1)
    sub_category = sheet[f"{cl}{rn}"].value
    sub_category_coordinate = sheet[f"{cl}{rn}"].coordinate
    _, min_row = get_split_coordinate(sub_category_coordinate, 1)
    subcategory_element_response_create(scoring_round, min_row=min_row, max_row=(min_row + 1), sheet=sheet, to_category_info=to_category_info,
                                        sub_category=sub_category)

    info_to_subcat_to_cat.update({category_name: to_category_info})  # Aggregate info for each Subcat at CAREGORY
    category_list.append(info_to_subcat_to_cat)

    # RELATIONSHIP MANAGEMENT CATEGORY

    cl, rn = get_split_coordinate(coordinate, 14)
    category_name = sheet[f"{cl}{rn}"].value
    category_coordinate = sheet[f"{cl}{rn}"].coordinate
    info_to_subcat_to_cat = {}
    to_category_info = []  # list of all subcat wit element info

    # Subcategory Issue Management
    cl, rn = get_split_coordinate(category_coordinate, 1)
    sub_category = sheet[f"{cl}{rn}"].value
    sub_category_coordinate = sheet[f"{cl}{rn}"].coordinate
    _, min_row = get_split_coordinate(sub_category_coordinate, 1)
    subcategory_element_response_create(scoring_round, min_row=min_row, max_row=(min_row + 4), sheet=sheet, to_category_info=to_category_info,
                                        sub_category=sub_category)

    # Subcategory Plan Management
    cl, rn = get_split_coordinate(category_coordinate, 9)
    sub_category = sheet[f"{cl}{rn}"].value
    sub_category_coordinate = sheet[f"{cl}{rn}"].coordinate
    _, min_row = get_split_coordinate(sub_category_coordinate, 1)
    subcategory_element_response_create(scoring_round, min_row=min_row, max_row=(min_row + 7), sheet=sheet, to_category_info=to_category_info,
                                        sub_category=sub_category)

    info_to_subcat_to_cat.update({category_name: to_category_info})  # Aggregate info for each Subcat at CAREGORY
    category_list.append(info_to_subcat_to_cat)

    # RISK MANAGEMENT CATEGORY

    cl, rn = get_split_coordinate(coordinate, 34)
    category_name = sheet[f"{cl}{rn}"].value
    category_coordinate = sheet[f"{cl}{rn}"].coordinate
    info_to_subcat_to_cat = {}
    to_category_info = []  # list of all subcat wit element info

    # Subcategory Assessment
    cl, rn = get_split_coordinate(category_coordinate, 1)
    sub_category = sheet[f"{cl}{rn}"].value
    sub_category_coordinate = sheet[f"{cl}{rn}"].coordinate
    _, min_row = get_split_coordinate(sub_category_coordinate, 1)
    subcategory_element_response_create(scoring_round, min_row=min_row, max_row=(min_row + 2), sheet=sheet, to_category_info=to_category_info,
                                        sub_category=sub_category)

    # Subcategory Mitigation Planning
    cl, rn = get_split_coordinate(category_coordinate, 7)
    sub_category = sheet[f"{cl}{rn}"].value
    sub_category_coordinate = sheet[f"{cl}{rn}"].coordinate
    _, min_row = get_split_coordinate(sub_category_coordinate, 1)
    subcategory_element_response_create(scoring_round, min_row=min_row, max_row=(min_row + 3), sheet=sheet, to_category_info=to_category_info,
                                        sub_category=sub_category)

    # Subcategory Model Definition
    cl, rn = get_split_coordinate(category_coordinate, 14)
    sub_category = sheet[f"{cl}{rn}"].value
    sub_category_coordinate = sheet[f"{cl}{rn}"].coordinate
    _, min_row = get_split_coordinate(sub_category_coordinate, 1)
    subcategory_element_response_create(scoring_round, min_row=min_row, max_row=(min_row + 3), sheet=sheet, to_category_info=to_category_info,
                                        sub_category=sub_category)

    # Subcategory Monitoring & Identification
    cl, rn = get_split_coordinate(category_coordinate, 21)
    sub_category = sheet[f"{cl}{rn}"].value
    sub_category_coordinate = sheet[f"{cl}{rn}"].coordinate
    _, min_row = get_split_coordinate(sub_category_coordinate, 1)
    subcategory_element_response_create(scoring_round, min_row=min_row, max_row=(min_row + 9), sheet=sheet, to_category_info=to_category_info,
                                        sub_category=sub_category)

    # Subcategory Regulatory Compliance
    cl, rn = get_split_coordinate(category_coordinate, 34)
    sub_category = sheet[f"{cl}{rn}"].value
    sub_category_coordinate = sheet[f"{cl}{rn}"].coordinate
    _, min_row = get_split_coordinate(sub_category_coordinate, 1)
    subcategory_element_response_create(scoring_round, min_row=min_row, max_row=(min_row + 7), sheet=sheet, to_category_info=to_category_info,
                                        sub_category=sub_category)

    # Subcategory Supplier Risk Management
    cl, rn = get_split_coordinate(category_coordinate, 45)
    sub_category = sheet[f"{cl}{rn}"].value
    sub_category_coordinate = sheet[f"{cl}{rn}"].coordinate
    _, min_row = get_split_coordinate(sub_category_coordinate, 1)
    subcategory_element_response_create(scoring_round, min_row=min_row, max_row=min_row, sheet=sheet, to_category_info=to_category_info,
                                        sub_category=sub_category)

    info_to_subcat_to_cat.update({category_name: to_category_info})  # Aggregate info for each Subcat at CAREGORY
    category_list.append(info_to_subcat_to_cat)

    # SUPPLIER INFORMATION MANAGEMENT CATEGORY

    cl, rn = get_split_coordinate(coordinate, 83)
    category_name = sheet[f"{cl}{rn}"].value
    category_coordinate = sheet[f"{cl}{rn}"].coordinate
    info_to_subcat_to_cat = {}
    to_category_info = []  # list of all subcat wit element info

    # Subcategory Discovery
    cl, rn = get_split_coordinate(category_coordinate, 1)
    sub_category = sheet[f"{cl}{rn}"].value
    sub_category_coordinate = sheet[f"{cl}{rn}"].coordinate
    _, min_row = get_split_coordinate(sub_category_coordinate, 1)
    subcategory_element_response_create(scoring_round, min_row=min_row, max_row=(min_row + 4), sheet=sheet, to_category_info=to_category_info,
                                        sub_category=sub_category)

    # Subcategory Onboarding Support
    cl, rn = get_split_coordinate(category_coordinate, 9)
    sub_category = sheet[f"{cl}{rn}"].value
    sub_category_coordinate = sheet[f"{cl}{rn}"].coordinate
    _, min_row = get_split_coordinate(sub_category_coordinate, 1)
    subcategory_element_response_create(scoring_round, min_row=min_row, max_row=(min_row + 1), sheet=sheet, to_category_info=to_category_info,
                                        sub_category=sub_category)

    # Subcategory Supply Base Profiling
    cl, rn = get_split_coordinate(category_coordinate, 14)
    sub_category = sheet[f"{cl}{rn}"].value
    sub_category_coordinate = sheet[f"{cl}{rn}"].coordinate
    _, min_row = get_split_coordinate(sub_category_coordinate, 1)
    subcategory_element_response_create(scoring_round, min_row=min_row, max_row=(min_row + 1), sheet=sheet, to_category_info=to_category_info,
                                        sub_category=sub_category)

    info_to_subcat_to_cat.update({category_name: to_category_info})  # Aggregate info for each Subcat at CAREGORY
    category_list.append(info_to_subcat_to_cat)

    # SUPPLIER PORTAL CATEGORY

    cl, rn = get_split_coordinate(coordinate, 102)
    category_name = sheet[f"{cl}{rn}"].value
    category_coordinate = sheet[f"{cl}{rn}"].coordinate
    info_to_subcat_to_cat = {}
    to_category_info = []  # list of all subcat wit element info

    # Subcategory Collaboration
    cl, rn = get_split_coordinate(category_coordinate, 1)
    sub_category = sheet[f"{cl}{rn}"].value
    sub_category_coordinate = sheet[f"{cl}{rn}"].coordinate
    _, min_row = get_split_coordinate(sub_category_coordinate, 1)
    subcategory_element_response_create(scoring_round, min_row=min_row, max_row=min_row, sheet=sheet, to_category_info=to_category_info,
                                        sub_category=sub_category)

    # Subcategory Information Management
    cl, rn = get_split_coordinate(category_coordinate, 5)
    sub_category = sheet[f"{cl}{rn}"].value
    sub_category_coordinate = sheet[f"{cl}{rn}"].coordinate
    _, min_row = get_split_coordinate(sub_category_coordinate, 1)
    subcategory_element_response_create(scoring_round, min_row=min_row, max_row=(min_row + 1), sheet=sheet, to_category_info=to_category_info,
                                        sub_category=sub_category)

    # Subcategory Performance Management
    cl, rn = get_split_coordinate(category_coordinate, 10)
    sub_category = sheet[f"{cl}{rn}"].value
    sub_category_coordinate = sheet[f"{cl}{rn}"].coordinate
    _, min_row = get_split_coordinate(sub_category_coordinate, 1)
    subcategory_element_response_create(scoring_round, min_row=min_row, max_row=min_row, sheet=sheet, to_category_info=to_category_info,
                                        sub_category=sub_category)

    # Subcategory Relationship Management
    cl, rn = get_split_coordinate(category_coordinate, 14)
    sub_category = sheet[f"{cl}{rn}"].value
    sub_category_coordinate = sheet[f"{cl}{rn}"].coordinate
    _, min_row = get_split_coordinate(sub_category_coordinate, 1)
    subcategory_element_response_create(scoring_round, min_row=min_row, max_row=(min_row + 2), sheet=sheet, to_category_info=to_category_info,
                                        sub_category=sub_category)
    # Subcategory Supplier Portal
    cl, rn = get_split_coordinate(category_coordinate, 20)
    sub_category = sheet[f"{cl}{rn}"].value
    sub_category_coordinate = sheet[f"{cl}{rn}"].coordinate
    _, min_row = get_split_coordinate(sub_category_coordinate, 1)
    subcategory_element_response_create(scoring_round, min_row=min_row, max_row=min_row, sheet=sheet, to_category_info=to_category_info,
                                        sub_category=sub_category)

    info_to_subcat_to_cat.update({category_name: to_category_info})  # Aggregate info for each Subcat at CAREGORY
    category_list.append(info_to_subcat_to_cat)

    pc_response.update({"Category": category_list})
    return pc_response


def services_response_create(workbook, coordinate, scoring_round):
    """
    Create response from SERVICES parsing
    :param file:
    :return:
    """
    sheet = workbook["RFI"]
    pc_response = {}  # crete response with all element information for each subcat for each category in PC
    pc_response.update({"Parent Category": sheet[coordinate].value})
    category_list = []  # list of all category in PC with subcat info and element data

    # S2P SERVICES CATEGORY
    cl, rn = get_split_coordinate(coordinate, 1)
    category_name = sheet[f"{cl}{rn}"].value
    category_coordinate = sheet[f"{cl}{rn}"].coordinate
    info_to_subcat_to_cat = {}
    to_category_info = []  # list of all subcat wit element info

    # Subcategory General
    sub_category = 'General'
    _, min_row = get_split_coordinate(category_coordinate, 1)
    subcategory_element_response_create(scoring_round, min_row=min_row, max_row=(min_row + 18), sheet=sheet, to_category_info=to_category_info,
                                        sub_category=sub_category)

    info_to_subcat_to_cat.update({category_name: to_category_info})  # Aggregate info for each Subcat at CAREGORY
    category_list.append(info_to_subcat_to_cat)

    # SPT SERVICES CATEGORY
    cl, rn = get_split_coordinate(coordinate, 23)
    category_name = sheet[f"{cl}{rn}"].value
    category_coordinate = sheet[f"{cl}{rn}"].coordinate
    info_to_subcat_to_cat = {}
    to_category_info = []  # list of all subcat wit element info

    # Subcategory General
    sub_category = 'General'
    _, min_row = get_split_coordinate(category_coordinate, 1)
    subcategory_element_response_create(scoring_round, min_row=min_row, max_row=min_row, sheet=sheet, to_category_info=to_category_info,
                                        sub_category=sub_category)

    info_to_subcat_to_cat.update({category_name: to_category_info})  # Aggregate info for each Subcat at CAREGORY
    category_list.append(info_to_subcat_to_cat)

    # CATEGORY Sourcing - SXM Services

    cl, rn = get_split_coordinate(coordinate, 27)
    category_name = sheet[f"{cl}{rn}"].value
    category_coordinate = sheet[f"{cl}{rn}"].coordinate
    info_to_subcat_to_cat = {}
    to_category_info = []  # list of all subcat wit element info

    # Subcategory General
    sub_category = 'General'
    _, min_row = get_split_coordinate(category_coordinate, 1)
    subcategory_element_response_create(scoring_round, min_row=min_row, max_row=(min_row + 2), sheet=sheet, to_category_info=to_category_info,
                                        sub_category=sub_category)

    info_to_subcat_to_cat.update({category_name: to_category_info})  # Aggregate info for each Subcat at CAREGORY
    category_list.append(info_to_subcat_to_cat)

    pc_response.update({"Category": category_list})
    return pc_response


def sourcing_response_create(workbook, coordinate, scoring_round):
    """
    Create response from SOURCING parsing
    :param file:
    :return:
    """
    sheet = workbook["RFI"]
    pc_response = {}  # crete response with all element information for each subcat for each category in PC
    pc_response.update({"Parent Category": sheet[coordinate].value})
    category_list = []  # list of all category in PC with subcat info and element data

    # OPPORTUNITY CATEGORY

    cl, rn = get_split_coordinate(coordinate, 1)
    category_name = sheet[f"{cl}{rn}"].value
    category_coordinate = sheet[f"{cl}{rn}"].coordinate
    info_to_subcat_to_cat = {}
    to_category_info = []  # list of all subcat wit element info

    # Subcategory Category Analysis
    cl, rn = get_split_coordinate(category_coordinate, 1)
    sub_category = sheet[f"{cl}{rn}"].value
    sub_category_coordinate = sheet[f"{cl}{rn}"].coordinate
    _, min_row = get_split_coordinate(sub_category_coordinate, 1)
    subcategory_element_response_create(scoring_round, min_row=min_row, max_row=(min_row + 13), sheet=sheet, to_category_info=to_category_info,
                                        sub_category=sub_category)
    # Subcategory Should-Cost Modelling
    cl, rn = get_split_coordinate(category_coordinate, 18)
    sub_category = sheet[f"{cl}{rn}"].value
    sub_category_coordinate = sheet[f"{cl}{rn}"].coordinate
    _, min_row = get_split_coordinate(sub_category_coordinate, 1)
    subcategory_element_response_create(scoring_round, min_row=min_row, max_row=(min_row + 6), sheet=sheet, to_category_info=to_category_info,
                                        sub_category=sub_category)

    info_to_subcat_to_cat.update({category_name: to_category_info})  # Aggregate info for each Subcat at CAREGORY
    category_list.append(info_to_subcat_to_cat)

    #  RFX / SURVEYS CATEGORY

    cl, rn = get_split_coordinate(coordinate, 29)
    category_name = sheet[f"{cl}{rn}"].value
    category_coordinate = sheet[f"{cl}{rn}"].coordinate
    info_to_subcat_to_cat = {}
    to_category_info = []  # list of all subcat wit element info

    # Subcategory Construction
    cl, rn = get_split_coordinate(category_coordinate, 1)
    sub_category = sheet[f"{cl}{rn}"].value
    sub_category_coordinate = sheet[f"{cl}{rn}"].coordinate
    _, min_row = get_split_coordinate(sub_category_coordinate, 1)
    subcategory_element_response_create(scoring_round, min_row=min_row, max_row=(min_row + 20), sheet=sheet, to_category_info=to_category_info,
                                        sub_category=sub_category)
    # Subcategory Evaluation mechanisms
    cl, rn = get_split_coordinate(category_coordinate, 25)
    sub_category = sheet[f"{cl}{rn}"].value
    sub_category_coordinate = sheet[f"{cl}{rn}"].coordinate
    _, min_row = get_split_coordinate(sub_category_coordinate, 1)
    subcategory_element_response_create(scoring_round, min_row=min_row, max_row=(min_row + 3), sheet=sheet, to_category_info=to_category_info,
                                        sub_category=sub_category)
    # Subcategory RFX Management Capabilities
    cl, rn = get_split_coordinate(category_coordinate, 32)
    sub_category = sheet[f"{cl}{rn}"].value
    sub_category_coordinate = sheet[f"{cl}{rn}"].coordinate
    _, min_row = get_split_coordinate(sub_category_coordinate, 1)
    subcategory_element_response_create(scoring_round, min_row=min_row, max_row=(min_row + 5), sheet=sheet, to_category_info=to_category_info,
                                        sub_category=sub_category)
    # Subcategory Auction
    cl, rn = get_split_coordinate(category_coordinate, 41)
    sub_category = sheet[f"{cl}{rn}"].value
    sub_category_coordinate = sheet[f"{cl}{rn}"].coordinate
    _, min_row = get_split_coordinate(sub_category_coordinate, 1)
    subcategory_element_response_create(scoring_round, min_row=min_row, max_row=(min_row + 10), sheet=sheet, to_category_info=to_category_info,
                                        sub_category=sub_category)

    info_to_subcat_to_cat.update({category_name: to_category_info})  # Aggregate info for each Subcat at CAREGORY
    category_list.append(info_to_subcat_to_cat)

    # OPTIMIZATION CATEGORY

    cl, rn = get_split_coordinate(coordinate, 84)
    category_name = sheet[f"{cl}{rn}"].value
    category_coordinate = sheet[f"{cl}{rn}"].coordinate
    info_to_subcat_to_cat = {}
    to_category_info = []  # list of all subcat wit element info

    # Subcategory Foundations
    cl, rn = get_split_coordinate(category_coordinate, 1)
    sub_category = sheet[f"{cl}{rn}"].value
    sub_category_coordinate = sheet[f"{cl}{rn}"].coordinate
    _, min_row = get_split_coordinate(sub_category_coordinate, 1)
    subcategory_element_response_create(scoring_round, min_row=min_row, max_row=(min_row + 8), sheet=sheet, to_category_info=to_category_info,
                                        sub_category=sub_category)

    # Subcategory Sophisticated Constraint Analysis
    cl, rn = get_split_coordinate(category_coordinate, 13)
    sub_category = sheet[f"{cl}{rn}"].value
    sub_category_coordinate = sheet[f"{cl}{rn}"].coordinate
    _, min_row = get_split_coordinate(sub_category_coordinate, 1)
    subcategory_element_response_create(scoring_round, min_row=min_row, max_row=(min_row + 7), sheet=sheet, to_category_info=to_category_info,
                                        sub_category=sub_category)

    # Subcategory Freight Analytics
    cl, rn = get_split_coordinate(category_coordinate, 24)
    sub_category = sheet[f"{cl}{rn}"].value
    sub_category_coordinate = sheet[f"{cl}{rn}"].coordinate
    _, min_row = get_split_coordinate(sub_category_coordinate, 1)
    subcategory_element_response_create(scoring_round, min_row=min_row, max_row=(min_row + 4), sheet=sheet, to_category_info=to_category_info,
                                        sub_category=sub_category)

    info_to_subcat_to_cat.update({category_name: to_category_info})  # Aggregate info for each Subcat at CAREGORY
    category_list.append(info_to_subcat_to_cat)

    # PERFOMANCE MANAGEMENT CATEGORY

    cl, rn = get_split_coordinate(coordinate, 116)
    category_name = sheet[f"{cl}{rn}"].value
    category_coordinate = sheet[f"{cl}{rn}"].coordinate
    info_to_subcat_to_cat = {}
    to_category_info = []  # list of all subcat wit element info

    # Subcategory Core Capabilities
    cl, rn = get_split_coordinate(category_coordinate, 1)
    sub_category = sheet[f"{cl}{rn}"].value
    sub_category_coordinate = sheet[f"{cl}{rn}"].coordinate
    _, min_row = get_split_coordinate(sub_category_coordinate, 1)
    subcategory_element_response_create(scoring_round, min_row=min_row, max_row=(min_row + 7), sheet=sheet, to_category_info=to_category_info,
                                        sub_category=sub_category)

    info_to_subcat_to_cat.update({category_name: to_category_info})  # Aggregate info for each Subcat at CAREGORY
    category_list.append(info_to_subcat_to_cat)

    # CLM SUPPORT CATEGORY

    cl, rn = get_split_coordinate(coordinate, 128)
    category_name = sheet[f"{cl}{rn}"].value
    category_coordinate = sheet[f"{cl}{rn}"].coordinate
    info_to_subcat_to_cat = {}
    to_category_info = []  # list of all subcat wit element info

    # Subcategory Core Capabilities
    cl, rn = get_split_coordinate(category_coordinate, 1)
    sub_category = sheet[f"{cl}{rn}"].value
    sub_category_coordinate = sheet[f"{cl}{rn}"].coordinate
    _, min_row = get_split_coordinate(sub_category_coordinate, 1)
    subcategory_element_response_create(scoring_round, min_row=min_row, max_row=(min_row + 5), sheet=sheet, to_category_info=to_category_info,
                                        sub_category=sub_category)

    info_to_subcat_to_cat.update({category_name: to_category_info})  # Aggregate info for each Subcat at CAREGORY
    category_list.append(info_to_subcat_to_cat)

    pc_response.update({"Category": category_list})
    return pc_response


def sxm_response_create(workbook, coordinate, scoring_round):
    """
    Create response from SXM parsing
    :param file:
    :return:
    """
    sheet = workbook["RFI"]
    pc_response = {}  # crete response with all element information for each subcat for each category in PC
    pc_response.update({"Parent Category": sheet[coordinate].value})
    category_list = []  # list of all category in PC with subcat info and element data

    # EXTENDED SIM CATEGORY

    cl, rn = get_split_coordinate(coordinate, 1)
    category_name = sheet[f"{cl}{rn}"].value
    category_coordinate = sheet[f"{cl}{rn}"].coordinate
    info_to_subcat_to_cat = {}
    to_category_info = []  # list of all subcat wit element info

    # Subcategory General
    sub_category = 'General'
    _, min_row = get_split_coordinate(category_coordinate, 1)
    subcategory_element_response_create(scoring_round, min_row=min_row, max_row=(min_row + 20), sheet=sheet, to_category_info=to_category_info,
                                        sub_category=sub_category)

    info_to_subcat_to_cat.update({category_name: to_category_info})  # Aggregate info for each Subcat at CATEGORY
    category_list.append(info_to_subcat_to_cat)

    # SXM CATEGORY

    cl, rn = get_split_coordinate(coordinate, 25)
    category_name = sheet[f"{cl}{rn}"].value
    category_coordinate = sheet[f"{cl}{rn}"].coordinate
    info_to_subcat_to_cat = {}
    to_category_info = []  # list of all subcat wit element info

    # Subcategory Development & Innovation Management
    cl, rn = get_split_coordinate(category_coordinate, 1)
    sub_category = sheet[f"{cl}{rn}"].value
    sub_category_coordinate = sheet[f"{cl}{rn}"].coordinate
    _, min_row = get_split_coordinate(sub_category_coordinate, 1)
    subcategory_element_response_create(scoring_round, min_row=min_row, max_row=(min_row + 4), sheet=sheet, to_category_info=to_category_info,
                                        sub_category=sub_category)
    # Subcategory NPD / NPI
    cl, rn = get_split_coordinate(category_coordinate, 9)
    sub_category = sheet[f"{cl}{rn}"].value
    sub_category_coordinate = sheet[f"{cl}{rn}"].coordinate
    _, min_row = get_split_coordinate(sub_category_coordinate, 1)
    subcategory_element_response_create(scoring_round, min_row=min_row, max_row=(min_row + 4), sheet=sheet, to_category_info=to_category_info,
                                        sub_category=sub_category)
    # Subcategory Out-of-the-Box Reporting
    # Subcategory NPD / NPI
    cl, rn = get_split_coordinate(category_coordinate, 17)
    sub_category = sheet[f"{cl}{rn}"].value
    sub_category_coordinate = sheet[f"{cl}{rn}"].coordinate
    _, min_row = get_split_coordinate(sub_category_coordinate, 1)
    subcategory_element_response_create(scoring_round, min_row=min_row, max_row=(min_row + 3), sheet=sheet, to_category_info=to_category_info,
                                        sub_category=sub_category)

    info_to_subcat_to_cat.update({category_name: to_category_info})  # Aggregate info for each Subcat at CATEGORY
    category_list.append(info_to_subcat_to_cat)

    pc_response.update({"Category": category_list})
    return pc_response


def spend_analytics_response_create(workbook, coordinate, scoring_round):
    """
    Create response from Spend Analytics parsing
    :param file:
    :return:
    """
    sheet = workbook["RFI"]
    pc_response = {}  # crete response with all element information for each subcat for each category in PC
    pc_response.update({"Parent Category": sheet[coordinate].value})
    category_list = []  # list of all category in PC with subcat info and element data

    # PROCESS SUPPORT CATEGORY

    cl, rn = get_split_coordinate(coordinate, 1)
    category_name = sheet[f"{cl}{rn}"].value
    category_coordinate = sheet[f"{cl}{rn}"].coordinate
    info_to_subcat_to_cat = {}
    to_category_info = []  # list of all subcat wit element info

    sub_category = 'General'
    _, min_row = get_split_coordinate(category_coordinate, 1)
    subcategory_element_response_create(scoring_round, min_row=min_row, max_row=(min_row + 17), sheet=sheet, to_category_info=to_category_info,
                                        sub_category=sub_category)

    info_to_subcat_to_cat.update({category_name: to_category_info})  # Aggregate info for each Subcat at CATEGORY
    category_list.append(info_to_subcat_to_cat)

    # FUNCTION SUPPORT CATEGORY

    cl, rn = get_split_coordinate(coordinate, 22)
    category_name = sheet[f"{cl}{rn}"].value
    category_coordinate = sheet[f"{cl}{rn}"].coordinate
    info_to_subcat_to_cat = {}
    to_category_info = []  # list of all subcat wit element info

    # Subcategory Raw Capability
    cl, rn = get_split_coordinate(category_coordinate, 1)
    sub_category = sheet[f"{cl}{rn}"].value
    sub_category_coordinate = sheet[f"{cl}{rn}"].coordinate
    _, min_row = get_split_coordinate(sub_category_coordinate, 1)
    subcategory_element_response_create(scoring_round, min_row=min_row, max_row=(min_row + 7), sheet=sheet, to_category_info=to_category_info,
                                        sub_category=sub_category)

    # Subcategory Out of the Box
    cl, rn = get_split_coordinate(category_coordinate, 12)
    sub_category = sheet[f"{cl}{rn}"].value
    sub_category_coordinate = sheet[f"{cl}{rn}"].coordinate
    _, min_row = get_split_coordinate(sub_category_coordinate, 1)
    subcategory_element_response_create(scoring_round, min_row=min_row, max_row=(min_row + 10), sheet=sheet, to_category_info=to_category_info,
                                        sub_category=sub_category)

    info_to_subcat_to_cat.update({category_name: to_category_info})  # Aggregate info for each Subcat at CATEGORY
    category_list.append(info_to_subcat_to_cat)

    pc_response.update({"Category": category_list})
    return pc_response


def clm_response_create(workbook, coordinate, scoring_round):
    """
    Create response from CLM parsing
    :param file:
    :return:
    """
    sheet = workbook["RFI"]
    pc_response = {}  # crete response with all element information for each subcat for each category in PC
    pc_response.update({"Parent Category": sheet[coordinate].value})
    category_list = []  # list of all category in PC with subcat info and element data

    # CONTACT INFORMATION MANAGEMENT CATEGORY

    cl, rn = get_split_coordinate(coordinate, 1)
    category_name = sheet[f"{cl}{rn}"].value
    category_coordinate = sheet[f"{cl}{rn}"].coordinate
    info_to_subcat_to_cat = {}
    to_category_info = []  # list of all subcat wit element info

    # Subcategory Core Contract modeling
    cl, rn = get_split_coordinate(category_coordinate, 1)
    sub_category = sheet[f"{cl}{rn}"].value
    sub_category_coordinate = sheet[f"{cl}{rn}"].coordinate
    _, min_row = get_split_coordinate(sub_category_coordinate, 1)
    subcategory_element_response_create(scoring_round, min_row=min_row, max_row=(min_row + 8), sheet=sheet, to_category_info=to_category_info,
                                        sub_category=sub_category)

    # # Subcategory Extended Contract Modeling and Analytics
    cl, rn = get_split_coordinate(category_coordinate, 13)
    sub_category = sheet[f"{cl}{rn}"].value
    sub_category_coordinate = sheet[f"{cl}{rn}"].coordinate
    _, min_row = get_split_coordinate(sub_category_coordinate, 1)
    subcategory_element_response_create(scoring_round, min_row=min_row, max_row=(min_row + 8), sheet=sheet, to_category_info=to_category_info,
                                        sub_category=sub_category)

    info_to_subcat_to_cat.update({category_name: to_category_info})  # Aggregate info for each Subcat at CAREGORY
    category_list.append(info_to_subcat_to_cat)

    # CONTACT PROCESS MANAGEMENT CATEGORY

    cl, rn = get_split_coordinate(coordinate, 26)
    category_name = sheet[f"{cl}{rn}"].value
    category_coordinate = sheet[f"{cl}{rn}"].coordinate
    info_to_subcat_to_cat = {}
    to_category_info = []  # list of all subcat wit element info

    # Subcategory Contract Expiry & Renewal Management
    cl, rn = get_split_coordinate(category_coordinate, 1)
    sub_category = sheet[f"{cl}{rn}"].value
    sub_category_coordinate = sheet[f"{cl}{rn}"].coordinate
    _, min_row = get_split_coordinate(sub_category_coordinate, 1)
    subcategory_element_response_create(scoring_round, min_row=min_row, max_row=(min_row + 2), sheet=sheet, to_category_info=to_category_info,
                                        sub_category=sub_category)

    # Subcategory Contract Creation and Authoring

    cl, rn = get_split_coordinate(category_coordinate, 7)
    sub_category = sheet[f"{cl}{rn}"].value
    sub_category_coordinate = sheet[f"{cl}{rn}"].coordinate
    _, min_row = get_split_coordinate(sub_category_coordinate, 1)
    subcategory_element_response_create(scoring_round, min_row=min_row, max_row=(min_row + 5), sheet=sheet, to_category_info=to_category_info,
                                        sub_category=sub_category)

    # Subcategory Contract Collaboration
    cl, rn = get_split_coordinate(category_coordinate, 16)
    sub_category = sheet[f"{cl}{rn}"].value
    sub_category_coordinate = sheet[f"{cl}{rn}"].coordinate
    _, min_row = get_split_coordinate(sub_category_coordinate, 1)
    subcategory_element_response_create(scoring_round, min_row=min_row, max_row=(min_row + 3), sheet=sheet, to_category_info=to_category_info,
                                        sub_category=sub_category)

    # Subcategory Contract Performance Management
    cl, rn = get_split_coordinate(category_coordinate, 23)
    sub_category = sheet[f"{cl}{rn}"].value
    sub_category_coordinate = sheet[f"{cl}{rn}"].coordinate
    _, min_row = get_split_coordinate(sub_category_coordinate, 1)
    subcategory_element_response_create(scoring_round, min_row=min_row, max_row=(min_row + 2), sheet=sheet, to_category_info=to_category_info,
                                        sub_category=sub_category)

    info_to_subcat_to_cat.update({category_name: to_category_info})  # Aggregate info for each Subcat at CAREGORY
    category_list.append(info_to_subcat_to_cat)

    # ANALYTICS CATEGORY

    cl, rn = get_split_coordinate(coordinate, 55)
    category_name = sheet[f"{cl}{rn}"].value
    category_coordinate = sheet[f"{cl}{rn}"].coordinate
    info_to_subcat_to_cat = {}
    to_category_info = []  # list of all subcat wit element info

    # Subcategory Performance Management Analytics
    cl, rn = get_split_coordinate(category_coordinate, 1)
    sub_category = sheet[f"{cl}{rn}"].value
    sub_category_coordinate = sheet[f"{cl}{rn}"].coordinate
    _, min_row = get_split_coordinate(sub_category_coordinate, 1)
    subcategory_element_response_create(scoring_round, min_row=min_row, max_row=(min_row + 1), sheet=sheet, to_category_info=to_category_info,
                                        sub_category=sub_category)

    # Subcategory Knowledge Management and Expertise
    cl, rn = get_split_coordinate(category_coordinate, 6)
    sub_category = sheet[f"{cl}{rn}"].value
    sub_category_coordinate = sheet[f"{cl}{rn}"].coordinate
    _, min_row = get_split_coordinate(sub_category_coordinate, 1)
    subcategory_element_response_create(scoring_round, min_row=min_row, max_row=(min_row + 2), sheet=sheet, to_category_info=to_category_info,
                                        sub_category=sub_category)

    info_to_subcat_to_cat.update({category_name: to_category_info})  # Aggregate info for each Subcat at CAREGORY
    category_list.append(info_to_subcat_to_cat)

    pc_response.update({"Category": category_list})
    return pc_response


def eprocurement_response_create(workbook, coordinate, scoring_round):
    """
    Create response from eProcurement parsing
    :param file:
    :return:
    """
    sheet = workbook["RFI"]
    pc_response = {}  # crete response with all element information for each subcat for each category in PC
    pc_response.update({"Parent Category": sheet[coordinate].value})
    category_list = []  # list of all category in PC with subcat info and element data

    # CATALOG MANAGEMENT CATEGORY

    cl, rn = get_split_coordinate(coordinate, 1)
    category_name = sheet[f"{cl}{rn}"].value
    category_coordinate = sheet[f"{cl}{rn}"].coordinate
    info_to_subcat_to_cat = {}
    to_category_info = []  # list of all subcat wit element info

    # Subcategory Catalog Creation / Onboarding
    cl, rn = get_split_coordinate(category_coordinate, 1)
    sub_category = sheet[f"{cl}{rn}"].value
    sub_category_coordinate = sheet[f"{cl}{rn}"].coordinate
    _, min_row = get_split_coordinate(sub_category_coordinate, 1)
    subcategory_element_response_create(scoring_round, min_row=min_row, max_row=(min_row + 9), sheet=sheet, to_category_info=to_category_info,
                                        sub_category=sub_category)

    # Subcategory Catalog Objects
    cl, rn = get_split_coordinate(category_coordinate, 14)
    sub_category = sheet[f"{cl}{rn}"].value
    sub_category_coordinate = sheet[f"{cl}{rn}"].coordinate
    _, min_row = get_split_coordinate(sub_category_coordinate, 1)
    subcategory_element_response_create(scoring_round, min_row=min_row, max_row=(min_row + 2), sheet=sheet, to_category_info=to_category_info,
                                        sub_category=sub_category)

    # Subcategory Catalog Data Quality Control
    cl, rn = get_split_coordinate(category_coordinate, 20)
    sub_category = sheet[f"{cl}{rn}"].value
    sub_category_coordinate = sheet[f"{cl}{rn}"].coordinate
    _, min_row = get_split_coordinate(sub_category_coordinate, 1)
    subcategory_element_response_create(scoring_round, min_row=min_row, max_row=(min_row + 5), sheet=sheet, to_category_info=to_category_info,
                                        sub_category=sub_category)

    # Subcategory Catalog Approvals  / Validations
    cl, rn = get_split_coordinate(category_coordinate, 29)
    sub_category = sheet[f"{cl}{rn}"].value
    sub_category_coordinate = sheet[f"{cl}{rn}"].coordinate
    _, min_row = get_split_coordinate(sub_category_coordinate, 1)
    subcategory_element_response_create(scoring_round, min_row=min_row, max_row=min_row, sheet=sheet, to_category_info=to_category_info,
                                        sub_category=sub_category)

    # Subcategory Catalog Maintenance
    cl, rn = get_split_coordinate(category_coordinate, 33)
    sub_category = sheet[f"{cl}{rn}"].value
    sub_category_coordinate = sheet[f"{cl}{rn}"].coordinate
    _, min_row = get_split_coordinate(sub_category_coordinate, 1)
    subcategory_element_response_create(scoring_round, min_row=min_row, max_row=min_row, sheet=sheet, to_category_info=to_category_info,
                                        sub_category=sub_category)

    # Subcategory Catalog Mobility
    cl, rn = get_split_coordinate(category_coordinate, 37)
    sub_category = sheet[f"{cl}{rn}"].value
    sub_category_coordinate = sheet[f"{cl}{rn}"].coordinate
    _, min_row = get_split_coordinate(sub_category_coordinate, 1)
    subcategory_element_response_create(scoring_round, min_row=min_row, max_row=min_row, sheet=sheet, to_category_info=to_category_info,
                                        sub_category=sub_category)

    # Subcategory Catalog Analytics
    cl, rn = get_split_coordinate(category_coordinate, 41)
    sub_category = sheet[f"{cl}{rn}"].value
    sub_category_coordinate = sheet[f"{cl}{rn}"].coordinate
    _, min_row = get_split_coordinate(sub_category_coordinate, 1)
    subcategory_element_response_create(scoring_round, min_row=min_row, max_row=min_row, sheet=sheet, to_category_info=to_category_info,
                                        sub_category=sub_category)

    # Subcategory Internet Shopping / Distributed Content
    cl, rn = get_split_coordinate(category_coordinate, 45)
    sub_category = sheet[f"{cl}{rn}"].value
    sub_category_coordinate = sheet[f"{cl}{rn}"].coordinate
    _, min_row = get_split_coordinate(sub_category_coordinate, 1)
    subcategory_element_response_create(scoring_round, min_row=min_row, max_row=min_row, sheet=sheet, to_category_info=to_category_info,
                                        sub_category=sub_category)

    # Subcategory Catalog Roadmap
    cl, rn = get_split_coordinate(category_coordinate, 49)
    sub_category = sheet[f"{cl}{rn}"].value
    sub_category_coordinate = sheet[f"{cl}{rn}"].coordinate
    _, min_row = get_split_coordinate(sub_category_coordinate, 1)
    subcategory_element_response_create(scoring_round, min_row=min_row, max_row=min_row, sheet=sheet, to_category_info=to_category_info,
                                        sub_category=sub_category)

    info_to_subcat_to_cat.update({category_name: to_category_info})
    category_list.append(info_to_subcat_to_cat)

    # CATEGORY Requisitioning
    cl, rn = get_split_coordinate(coordinate, 54)
    category_name = sheet[f"{cl}{rn}"].value
    category_coordinate = sheet[f"{cl}{rn}"].coordinate
    info_to_subcat_to_cat = {}
    to_category_info = []  # list of all subcat wit element info

    # Subcategory Requisitioning Setup
    cl, rn = get_split_coordinate(category_coordinate, 1)
    sub_category = sheet[f"{cl}{rn}"].value
    sub_category_coordinate = sheet[f"{cl}{rn}"].coordinate
    _, min_row = get_split_coordinate(sub_category_coordinate, 1)
    subcategory_element_response_create(scoring_round, min_row=min_row, max_row=(min_row + 3), sheet=sheet, to_category_info=to_category_info,
                                        sub_category=sub_category)

    # Subcategory Profiles Setup
    cl, rn = get_split_coordinate(category_coordinate, 8)
    sub_category = sheet[f"{cl}{rn}"].value
    sub_category_coordinate = sheet[f"{cl}{rn}"].coordinate
    _, min_row = get_split_coordinate(sub_category_coordinate, 1)
    subcategory_element_response_create(scoring_round, min_row=min_row, max_row=(min_row + 1), sheet=sheet, to_category_info=to_category_info,
                                        sub_category=sub_category)

    # Subcategory Marketplace User Interface
    cl, rn = get_split_coordinate(category_coordinate, 13)
    sub_category = sheet[f"{cl}{rn}"].value
    sub_category_coordinate = sheet[f"{cl}{rn}"].coordinate
    _, min_row = get_split_coordinate(sub_category_coordinate, 1)
    subcategory_element_response_create(scoring_round, min_row=min_row, max_row=(min_row + 2), sheet=sheet, to_category_info=to_category_info,
                                        sub_category=sub_category)

    # Subcategory Search Engine
    cl, rn = get_split_coordinate(category_coordinate, 19)
    sub_category = sheet[f"{cl}{rn}"].value
    sub_category_coordinate = sheet[f"{cl}{rn}"].coordinate
    _, min_row = get_split_coordinate(sub_category_coordinate, 1)
    subcategory_element_response_create(scoring_round, min_row=min_row, max_row=(min_row + 6), sheet=sheet, to_category_info=to_category_info,
                                        sub_category=sub_category)

    # Subcategory Third-Party Content
    cl, rn = get_split_coordinate(category_coordinate, 29)
    sub_category = sheet[f"{cl}{rn}"].value
    sub_category_coordinate = sheet[f"{cl}{rn}"].coordinate
    _, min_row = get_split_coordinate(sub_category_coordinate, 1)
    subcategory_element_response_create(scoring_round, min_row=min_row, max_row=(min_row + 3), sheet=sheet, to_category_info=to_category_info,
                                        sub_category=sub_category)

    # Subcategory Requisitioning Process
    cl, rn = get_split_coordinate(category_coordinate, 36)
    sub_category = sheet[f"{cl}{rn}"].value
    sub_category_coordinate = sheet[f"{cl}{rn}"].coordinate
    _, min_row = get_split_coordinate(sub_category_coordinate, 1)
    subcategory_element_response_create(scoring_round, min_row=min_row, max_row=(min_row + 11), sheet=sheet, to_category_info=to_category_info,
                                        sub_category=sub_category)

    # Subcategory Sourcing Integration
    cl, rn = get_split_coordinate(category_coordinate, 51)
    sub_category = sheet[f"{cl}{rn}"].value
    sub_category_coordinate = sheet[f"{cl}{rn}"].coordinate
    _, min_row = get_split_coordinate(sub_category_coordinate, 1)
    subcategory_element_response_create(scoring_round, min_row=min_row, max_row=(min_row + 2), sheet=sheet, to_category_info=to_category_info,
                                        sub_category=sub_category)

    # Subcategory Direct Material Requisitioning
    cl, rn = get_split_coordinate(category_coordinate, 57)
    sub_category = sheet[f"{cl}{rn}"].value
    sub_category_coordinate = sheet[f"{cl}{rn}"].coordinate
    _, min_row = get_split_coordinate(sub_category_coordinate, 1)
    subcategory_element_response_create(scoring_round, min_row=min_row, max_row=(min_row + 4), sheet=sheet, to_category_info=to_category_info,
                                        sub_category=sub_category)

    # Subcategory Guided Buying
    cl, rn = get_split_coordinate(category_coordinate, 65)
    sub_category = sheet[f"{cl}{rn}"].value
    sub_category_coordinate = sheet[f"{cl}{rn}"].coordinate
    _, min_row = get_split_coordinate(sub_category_coordinate, 1)
    subcategory_element_response_create(scoring_round, min_row=min_row, max_row=(min_row + 8), sheet=sheet, to_category_info=to_category_info,
                                        sub_category=sub_category)

    # Subcategory Requisitioning Help & Support
    cl, rn = get_split_coordinate(category_coordinate, 77)
    sub_category = sheet[f"{cl}{rn}"].value
    sub_category_coordinate = sheet[f"{cl}{rn}"].coordinate
    _, min_row = get_split_coordinate(sub_category_coordinate, 1)
    subcategory_element_response_create(scoring_round, min_row=min_row, max_row=(min_row + 2), sheet=sheet, to_category_info=to_category_info,
                                        sub_category=sub_category)

    # Subcategory Shopping Cart / Checkout Process
    cl, rn = get_split_coordinate(category_coordinate, 83)
    sub_category = sheet[f"{cl}{rn}"].value
    sub_category_coordinate = sheet[f"{cl}{rn}"].coordinate
    _, min_row = get_split_coordinate(sub_category_coordinate, 1)
    subcategory_element_response_create(scoring_round, min_row=min_row, max_row=(min_row + 7), sheet=sheet, to_category_info=to_category_info,
                                        sub_category=sub_category)

    # Subcategory Requisitioning Budget Checking Process
    cl, rn = get_split_coordinate(category_coordinate, 94)
    sub_category = sheet[f"{cl}{rn}"].value
    sub_category_coordinate = sheet[f"{cl}{rn}"].coordinate
    _, min_row = get_split_coordinate(sub_category_coordinate, 1)
    subcategory_element_response_create(scoring_round, min_row=min_row, max_row=min_row, sheet=sheet, to_category_info=to_category_info,
                                        sub_category=sub_category)

    # Requisitioning Inventory Checking Process
    cl, rn = get_split_coordinate(category_coordinate, 98)
    sub_category = sheet[f"{cl}{rn}"].value
    sub_category_coordinate = sheet[f"{cl}{rn}"].coordinate
    _, min_row = get_split_coordinate(sub_category_coordinate, 1)
    subcategory_element_response_create(scoring_round, min_row=min_row, max_row=(min_row + 1), sheet=sheet, to_category_info=to_category_info,
                                        sub_category=sub_category)

    # Approval Process / Approval Engine
    cl, rn = get_split_coordinate(category_coordinate, 103)
    sub_category = sheet[f"{cl}{rn}"].value
    sub_category_coordinate = sheet[f"{cl}{rn}"].coordinate
    _, min_row = get_split_coordinate(sub_category_coordinate, 1)
    subcategory_element_response_create(scoring_round, min_row=min_row, max_row=(min_row + 3), sheet=sheet, to_category_info=to_category_info,
                                        sub_category=sub_category)

    # Mobility
    cl, rn = get_split_coordinate(category_coordinate, 110)
    sub_category = sheet[f"{cl}{rn}"].value
    sub_category_coordinate = sheet[f"{cl}{rn}"].coordinate
    _, min_row = get_split_coordinate(sub_category_coordinate, 1)
    subcategory_element_response_create(scoring_round, min_row=min_row, max_row=min_row, sheet=sheet, to_category_info=to_category_info,
                                        sub_category=sub_category)

    # Requisitioning Analytics
    cl, rn = get_split_coordinate(category_coordinate, 114)
    sub_category = sheet[f"{cl}{rn}"].value
    sub_category_coordinate = sheet[f"{cl}{rn}"].coordinate
    _, min_row = get_split_coordinate(sub_category_coordinate, 1)
    subcategory_element_response_create(scoring_round, min_row=min_row, max_row=min_row, sheet=sheet, to_category_info=to_category_info,
                                        sub_category=sub_category)

    # Requisition Roadmap
    cl, rn = get_split_coordinate(category_coordinate, 118)
    sub_category = sheet[f"{cl}{rn}"].value
    sub_category_coordinate = sheet[f"{cl}{rn}"].coordinate
    _, min_row = get_split_coordinate(sub_category_coordinate, 1)
    subcategory_element_response_create(scoring_round, min_row=min_row, max_row=min_row, sheet=sheet, to_category_info=to_category_info,
                                        sub_category=sub_category)

    info_to_subcat_to_cat.update({category_name: to_category_info})
    category_list.append(info_to_subcat_to_cat)

    # CATEGORY Ordering
    cl, rn = get_split_coordinate(coordinate, 176)
    category_name = sheet[f"{cl}{rn}"].value
    category_coordinate = sheet[f"{cl}{rn}"].coordinate
    info_to_subcat_to_cat = {}
    to_category_info = []  # list of all subcat wit element info

    # Subcategory Order Setup
    cl, rn = get_split_coordinate(category_coordinate, 1)
    sub_category = sheet[f"{cl}{rn}"].value
    sub_category_coordinate = sheet[f"{cl}{rn}"].coordinate
    _, min_row = get_split_coordinate(sub_category_coordinate, 1)
    subcategory_element_response_create(scoring_round, min_row=min_row, max_row=(min_row + 1), sheet=sheet, to_category_info=to_category_info,
                                        sub_category=sub_category)

    # Order Creation
    cl, rn = get_split_coordinate(category_coordinate, 6)
    sub_category = sheet[f"{cl}{rn}"].value
    sub_category_coordinate = sheet[f"{cl}{rn}"].coordinate
    _, min_row = get_split_coordinate(sub_category_coordinate, 1)
    subcategory_element_response_create(scoring_round, min_row=min_row, max_row=(min_row + 7), sheet=sheet, to_category_info=to_category_info,
                                        sub_category=sub_category)

    # Contract Compliance
    cl, rn = get_split_coordinate(category_coordinate, 17)
    sub_category = sheet[f"{cl}{rn}"].value
    sub_category_coordinate = sheet[f"{cl}{rn}"].coordinate
    _, min_row = get_split_coordinate(sub_category_coordinate, 1)
    subcategory_element_response_create(scoring_round, min_row=min_row, max_row=min_row, sheet=sheet, to_category_info=to_category_info,
                                        sub_category=sub_category)

    # Order Processing (buy-side)
    cl, rn = get_split_coordinate(category_coordinate, 21)
    sub_category = sheet[f"{cl}{rn}"].value
    sub_category_coordinate = sheet[f"{cl}{rn}"].coordinate
    _, min_row = get_split_coordinate(sub_category_coordinate, 1)
    subcategory_element_response_create(scoring_round, min_row=min_row, max_row=(min_row + 4), sheet=sheet, to_category_info=to_category_info,
                                        sub_category=sub_category)

    # Order Delivery / Communication
    cl, rn = get_split_coordinate(category_coordinate, 29)
    sub_category = sheet[f"{cl}{rn}"].value
    sub_category_coordinate = sheet[f"{cl}{rn}"].coordinate
    _, min_row = get_split_coordinate(sub_category_coordinate, 1)
    subcategory_element_response_create(scoring_round, min_row=min_row, max_row=min_row, sheet=sheet, to_category_info=to_category_info,
                                        sub_category=sub_category)

    # Order Collaboration (buyer/supplier)
    cl, rn = get_split_coordinate(category_coordinate, 33)
    sub_category = sheet[f"{cl}{rn}"].value
    sub_category_coordinate = sheet[f"{cl}{rn}"].coordinate
    _, min_row = get_split_coordinate(sub_category_coordinate, 1)
    subcategory_element_response_create(scoring_round, min_row=min_row, max_row=min_row, sheet=sheet, to_category_info=to_category_info,
                                        sub_category=sub_category)

    # Order Processing (supply-side)
    cl, rn = get_split_coordinate(category_coordinate, 37)
    sub_category = sheet[f"{cl}{rn}"].value
    sub_category_coordinate = sheet[f"{cl}{rn}"].coordinate
    _, min_row = get_split_coordinate(sub_category_coordinate, 1)
    subcategory_element_response_create(scoring_round, min_row=min_row, max_row=(min_row + 2), sheet=sheet, to_category_info=to_category_info,
                                        sub_category=sub_category)

    # PO Mobility
    cl, rn = get_split_coordinate(category_coordinate, 43)
    sub_category = sheet[f"{cl}{rn}"].value
    sub_category_coordinate = sheet[f"{cl}{rn}"].coordinate
    _, min_row = get_split_coordinate(sub_category_coordinate, 1)
    subcategory_element_response_create(scoring_round, min_row=min_row, max_row=min_row, sheet=sheet, to_category_info=to_category_info,
                                        sub_category=sub_category)

    # Ordering Analytics
    cl, rn = get_split_coordinate(category_coordinate, 47)
    sub_category = sheet[f"{cl}{rn}"].value
    sub_category_coordinate = sheet[f"{cl}{rn}"].coordinate
    _, min_row = get_split_coordinate(sub_category_coordinate, 1)
    subcategory_element_response_create(scoring_round, min_row=min_row, max_row=min_row, sheet=sheet, to_category_info=to_category_info,
                                        sub_category=sub_category)

    # PO Roadmap
    cl, rn = get_split_coordinate(category_coordinate, 51)
    sub_category = sheet[f"{cl}{rn}"].value
    sub_category_coordinate = sheet[f"{cl}{rn}"].coordinate
    _, min_row = get_split_coordinate(sub_category_coordinate, 1)
    subcategory_element_response_create(scoring_round, min_row=min_row, max_row=min_row, sheet=sheet, to_category_info=to_category_info,
                                        sub_category=sub_category)

    info_to_subcat_to_cat.update({category_name: to_category_info})
    category_list.append(info_to_subcat_to_cat)

    # CATEGORY Receiving
    cl, rn = get_split_coordinate(coordinate, 231)
    category_name = sheet[f"{cl}{rn}"].value
    category_coordinate = sheet[f"{cl}{rn}"].coordinate
    info_to_subcat_to_cat = {}
    to_category_info = []  # list of all subcat wit element info

    # Subcategory Receiving Setup
    cl, rn = get_split_coordinate(category_coordinate, 1)
    sub_category = sheet[f"{cl}{rn}"].value
    sub_category_coordinate = sheet[f"{cl}{rn}"].coordinate
    _, min_row = get_split_coordinate(sub_category_coordinate, 1)
    subcategory_element_response_create(scoring_round, min_row=min_row, max_row=min_row, sheet=sheet, to_category_info=to_category_info,
                                        sub_category=sub_category)

    # Fulfillment
    cl, rn = get_split_coordinate(category_coordinate, 5)
    sub_category = sheet[f"{cl}{rn}"].value
    sub_category_coordinate = sheet[f"{cl}{rn}"].coordinate
    _, min_row = get_split_coordinate(sub_category_coordinate, 1)
    subcategory_element_response_create(scoring_round, min_row=min_row, max_row=(min_row + 1), sheet=sheet, to_category_info=to_category_info,
                                        sub_category=sub_category)

    # Receiving Process
    cl, rn = get_split_coordinate(category_coordinate, 10)
    sub_category = sheet[f"{cl}{rn}"].value
    sub_category_coordinate = sheet[f"{cl}{rn}"].coordinate
    _, min_row = get_split_coordinate(sub_category_coordinate, 1)
    subcategory_element_response_create(scoring_round, min_row=min_row, max_row=(min_row + 5), sheet=sheet, to_category_info=to_category_info,
                                        sub_category=sub_category)

    # Receiving Mobility
    cl, rn = get_split_coordinate(category_coordinate, 19)
    sub_category = sheet[f"{cl}{rn}"].value
    sub_category_coordinate = sheet[f"{cl}{rn}"].coordinate
    _, min_row = get_split_coordinate(sub_category_coordinate, 1)
    subcategory_element_response_create(scoring_round, min_row=min_row, max_row=min_row, sheet=sheet, to_category_info=to_category_info,
                                        sub_category=sub_category)

    # Receiving Analytics
    cl, rn = get_split_coordinate(category_coordinate, 23)
    sub_category = sheet[f"{cl}{rn}"].value
    sub_category_coordinate = sheet[f"{cl}{rn}"].coordinate
    _, min_row = get_split_coordinate(sub_category_coordinate, 1)
    subcategory_element_response_create(scoring_round, min_row=min_row, max_row=min_row, sheet=sheet,
                                        to_category_info=to_category_info,
                                        sub_category=sub_category)

    # Receiving Roadmap
    cl, rn = get_split_coordinate(category_coordinate, 27)
    sub_category = sheet[f"{cl}{rn}"].value
    sub_category_coordinate = sheet[f"{cl}{rn}"].coordinate
    _, min_row = get_split_coordinate(sub_category_coordinate, 1)
    subcategory_element_response_create(scoring_round, min_row=min_row, max_row=min_row, sheet=sheet,
                                        to_category_info=to_category_info,
                                        sub_category=sub_category)

    info_to_subcat_to_cat.update({category_name: to_category_info})
    category_list.append(info_to_subcat_to_cat)


    pc_response.update({"Category": category_list})
    return pc_response


def i2p_response_create(workbook, coordinate, scoring_round):
    """
    Create response from I2P parsing
    :param file:
    :return:
    """
    sheet = workbook["RFI"]
    pc_response = {}  # crete response with all element information for each subcat for each category in PC
    pc_response.update({"Parent Category": sheet[coordinate].value})
    category_list = []  # list of all category in PC with subcat info and element data

    # CATEGORY Invoicing

    cl, rn = get_split_coordinate(coordinate, 1)
    category_name = sheet[f"{cl}{rn}"].value
    category_coordinate = sheet[f"{cl}{rn}"].coordinate
    info_to_subcat_to_cat = {}
    to_category_info = []  # list of all subcat wit element info

    # Subcategory Catalog Creation / Onboarding Invoicing Setup
    cl, rn = get_split_coordinate(category_coordinate, 1)
    sub_category = sheet[f"{cl}{rn}"].value
    sub_category_coordinate = sheet[f"{cl}{rn}"].coordinate
    _, min_row = get_split_coordinate(sub_category_coordinate, 1)
    subcategory_element_response_create(scoring_round, min_row=min_row, max_row=(min_row + 2), sheet=sheet, to_category_info=to_category_info,
                                        sub_category=sub_category)

    # Subcategory Invoicing Creation / Capturing / Submission
    cl, rn = get_split_coordinate(category_coordinate, 7)
    sub_category = sheet[f"{cl}{rn}"].value
    sub_category_coordinate = sheet[f"{cl}{rn}"].coordinate
    _, min_row = get_split_coordinate(sub_category_coordinate, 1)
    subcategory_element_response_create(scoring_round, min_row=min_row, max_row=(min_row + 11), sheet=sheet, to_category_info=to_category_info,
                                        sub_category=sub_category)

    # Subcategory Services Invoicing & Contract Invoicing
    cl, rn = get_split_coordinate(category_coordinate, 22)
    sub_category = sheet[f"{cl}{rn}"].value
    sub_category_coordinate = sheet[f"{cl}{rn}"].coordinate
    _, min_row = get_split_coordinate(sub_category_coordinate, 1)
    subcategory_element_response_create(scoring_round, min_row=min_row, max_row=(min_row + 2), sheet=sheet, to_category_info=to_category_info,
                                        sub_category=sub_category)

    # Subcategory Invoice Compliance
    cl, rn = get_split_coordinate(category_coordinate, 28)
    sub_category = sheet[f"{cl}{rn}"].value
    sub_category_coordinate = sheet[f"{cl}{rn}"].coordinate
    _, min_row = get_split_coordinate(sub_category_coordinate, 1)
    subcategory_element_response_create(scoring_round, min_row=min_row, max_row=(min_row + 9), sheet=sheet, to_category_info=to_category_info,
                                        sub_category=sub_category)

    # Subcategory Invoice Validation / Approvals
    cl, rn = get_split_coordinate(category_coordinate, 41)
    sub_category = sheet[f"{cl}{rn}"].value
    sub_category_coordinate = sheet[f"{cl}{rn}"].coordinate
    _, min_row = get_split_coordinate(sub_category_coordinate, 1)
    subcategory_element_response_create(scoring_round, min_row=min_row, max_row=(min_row + 9), sheet=sheet, to_category_info=to_category_info,
                                        sub_category=sub_category)

    # Subcategory Invoice Collaboration
    cl, rn = get_split_coordinate(category_coordinate, 54)
    sub_category = sheet[f"{cl}{rn}"].value
    sub_category_coordinate = sheet[f"{cl}{rn}"].coordinate
    _, min_row = get_split_coordinate(sub_category_coordinate, 1)
    subcategory_element_response_create(scoring_round, min_row=min_row, max_row=(min_row + 3), sheet=sheet, to_category_info=to_category_info,
                                        sub_category=sub_category)

    # Subcategory Invoicing Integrations
    cl, rn = get_split_coordinate(category_coordinate, 61)
    sub_category = sheet[f"{cl}{rn}"].value
    sub_category_coordinate = sheet[f"{cl}{rn}"].coordinate
    _, min_row = get_split_coordinate(sub_category_coordinate, 1)
    subcategory_element_response_create(scoring_round, min_row=min_row, max_row=(min_row + 5), sheet=sheet, to_category_info=to_category_info,
                                        sub_category=sub_category)

    # Subcategory Invoicing Mobility
    cl, rn = get_split_coordinate(category_coordinate, 70)
    sub_category = sheet[f"{cl}{rn}"].value
    sub_category_coordinate = sheet[f"{cl}{rn}"].coordinate
    _, min_row = get_split_coordinate(sub_category_coordinate, 1)
    subcategory_element_response_create(scoring_round, min_row=min_row, max_row=min_row, sheet=sheet, to_category_info=to_category_info,
                                        sub_category=sub_category)

    # Subcategory Invoicing Analytics
    cl, rn = get_split_coordinate(category_coordinate, 74)
    sub_category = sheet[f"{cl}{rn}"].value
    sub_category_coordinate = sheet[f"{cl}{rn}"].coordinate
    _, min_row = get_split_coordinate(sub_category_coordinate, 1)
    subcategory_element_response_create(scoring_round, min_row=min_row, max_row=min_row, sheet=sheet, to_category_info=to_category_info,
                                        sub_category=sub_category)

    # Subcategory Invoicing Roadmap
    cl, rn = get_split_coordinate(category_coordinate, 78)
    sub_category = sheet[f"{cl}{rn}"].value
    sub_category_coordinate = sheet[f"{cl}{rn}"].coordinate
    _, min_row = get_split_coordinate(sub_category_coordinate, 1)
    subcategory_element_response_create(scoring_round, min_row=min_row, max_row=min_row, sheet=sheet, to_category_info=to_category_info,
                                        sub_category=sub_category)

    info_to_subcat_to_cat.update({category_name: to_category_info})  # Aggregate info for each Subcat at CAREGORY
    category_list.append(info_to_subcat_to_cat)

    # CATEGORY Payment & Financing

    cl, rn = get_split_coordinate(coordinate, 83)
    category_name = sheet[f"{cl}{rn}"].value
    category_coordinate = sheet[f"{cl}{rn}"].coordinate
    info_to_subcat_to_cat = {}
    to_category_info = []  # list of all subcat wit element info

    # Subcategory Payment Solution & Methods
    cl, rn = get_split_coordinate(category_coordinate, 1)
    sub_category = sheet[f"{cl}{rn}"].value
    sub_category_coordinate = sheet[f"{cl}{rn}"].coordinate
    _, min_row = get_split_coordinate(sub_category_coordinate, 1)
    subcategory_element_response_create(scoring_round, min_row=min_row, max_row=(min_row + 2), sheet=sheet, to_category_info=to_category_info,
                                        sub_category=sub_category)

    # Subcategory Payment Processing
    cl, rn = get_split_coordinate(category_coordinate, 7)
    sub_category = sheet[f"{cl}{rn}"].value
    sub_category_coordinate = sheet[f"{cl}{rn}"].coordinate
    _, min_row = get_split_coordinate(sub_category_coordinate, 1)
    subcategory_element_response_create(scoring_round, min_row=min_row, max_row=(min_row + 5), sheet=sheet, to_category_info=to_category_info,
                                        sub_category=sub_category)

    # Payment Cards
    cl, rn = get_split_coordinate(category_coordinate, 16)
    sub_category = sheet[f"{cl}{rn}"].value
    sub_category_coordinate = sheet[f"{cl}{rn}"].coordinate
    _, min_row = get_split_coordinate(sub_category_coordinate, 1)
    subcategory_element_response_create(scoring_round, min_row=min_row, max_row=(min_row + 3), sheet=sheet, to_category_info=to_category_info,
                                        sub_category=sub_category)

    # Early Payment Financing - Core
    cl, rn = get_split_coordinate(category_coordinate, 23)
    sub_category = sheet[f"{cl}{rn}"].value
    sub_category_coordinate = sheet[f"{cl}{rn}"].coordinate
    _, min_row = get_split_coordinate(sub_category_coordinate, 1)
    subcategory_element_response_create(scoring_round, min_row=min_row, max_row=(min_row + 15), sheet=sheet, to_category_info=to_category_info,
                                        sub_category=sub_category)

    # Financing Analytics
    cl, rn = get_split_coordinate(category_coordinate, 42)
    sub_category = sheet[f"{cl}{rn}"].value
    sub_category_coordinate = sheet[f"{cl}{rn}"].coordinate
    _, min_row = get_split_coordinate(sub_category_coordinate, 1)
    subcategory_element_response_create(scoring_round, min_row=min_row, max_row=min_row , sheet=sheet, to_category_info=to_category_info,
                                        sub_category=sub_category)

    # Payment & Financing Roadmap
    cl, rn = get_split_coordinate(category_coordinate, 46)
    sub_category = sheet[f"{cl}{rn}"].value
    sub_category_coordinate = sheet[f"{cl}{rn}"].coordinate
    _, min_row = get_split_coordinate(sub_category_coordinate, 1)
    subcategory_element_response_create(scoring_round, min_row=min_row, max_row=min_row, sheet=sheet, to_category_info=to_category_info,
                                        sub_category=sub_category)

    info_to_subcat_to_cat.update({category_name: to_category_info})  # Aggregate info for each Subcat at CAREGORY
    category_list.append(info_to_subcat_to_cat)

    # CATEGORY OPTIONAL For Specialized Personas(Additional coverage in SolutionMap)

    cl, rn = get_split_coordinate(coordinate, 133)
    category_name = sheet[f"{cl}{rn}"].value
    category_coordinate = sheet[f"{cl}{rn}"].coordinate
    info_to_subcat_to_cat = {}
    to_category_info = []  # list of all subcat wit element info

    # Subcategory Early Payment Financing - Specialized
    cl, rn = get_split_coordinate(category_coordinate, 1)
    sub_category = sheet[f"{cl}{rn}"].value
    sub_category_coordinate = sheet[f"{cl}{rn}"].coordinate
    _, min_row = get_split_coordinate(sub_category_coordinate, 1)
    subcategory_element_response_create(scoring_round, min_row=min_row, max_row=(min_row + 6), sheet=sheet, to_category_info=to_category_info,
                                        sub_category=sub_category)

    # Subcategory Dynamic Discounting (Specialized)
    cl, rn = get_split_coordinate(category_coordinate, 11)
    sub_category = sheet[f"{cl}{rn}"].value
    sub_category_coordinate = sheet[f"{cl}{rn}"].coordinate
    _, min_row = get_split_coordinate(sub_category_coordinate, 1)
    subcategory_element_response_create(scoring_round, min_row=min_row, max_row=(min_row + 5), sheet=sheet, to_category_info=to_category_info,
                                        sub_category=sub_category)

    # Supply Chain Finance (Specialized)
    cl, rn = get_split_coordinate(category_coordinate, 20)
    sub_category = sheet[f"{cl}{rn}"].value
    sub_category_coordinate = sheet[f"{cl}{rn}"].coordinate
    _, min_row = get_split_coordinate(sub_category_coordinate, 1)
    subcategory_element_response_create(scoring_round, min_row=min_row, max_row=(min_row + 9), sheet=sheet, to_category_info=to_category_info,
                                        sub_category=sub_category)

    info_to_subcat_to_cat.update({category_name: to_category_info})  # Aggregate info for each Subcat at CAREGORY
    category_list.append(info_to_subcat_to_cat)

    pc_response.update({"Category": category_list})
    return pc_response


pc_to_function_name = {
                            "COMMON S2P ": common_s2p_category_response_create,
                            "COMMON SOURCING – SXM ": common_sourcing_sxm_response_create,
                            "SERVICES ": services_response_create,
                            "SOURCING ": sourcing_response_create,
                            "SXM ": sxm_response_create,
                            "Spend Analytics ": spend_analytics_response_create,
                            "CLM ": clm_response_create,
                            "eProcurement ": eprocurement_response_create,
                            "I2P ": i2p_response_create,
                             }
# pc_to_function_name = {
#                             "COMMON S2P ": common_s2p_category_response_create,
#                             "COMMON SOURCING – SXM ": None,
#                             "SERVICES ": None,
#                             "SOURCING ": None,
#                             "SXM ": None,
#                             "Spend Analytics ": None,
#                             "CLM ": None,
#                             "eProcurement ": None,
#                             "I2P ": None,
#                              }