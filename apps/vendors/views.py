import csv
import os
import re
import string
import platform
import patoolib
from datetime import date
import pathlib
import shutil

from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill, Color, colors

from service.xml_file_upload_downlod import  get_excel_file_current_pc_for_parsing

from solution_project.settings.base import BASE_DIR
from django.core.files.storage import default_storage
from django.core.exceptions import ValidationError
from django.shortcuts import get_object_or_404
from django.http import HttpResponse
from django.db import transaction
from django.db.models import Max

from rest_framework.exceptions import ParseError
from rest_framework import generics
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework.parsers import MultiPartParser, FormParser, JSONParser
from rest_framework import status
from rest_framework import permissions

from service.csv_file_download import csv_file_parser, rfi_csv_file_parser
from service.xml_file_upload_downlod import InvalidFormatException, get_full_excel_file_response, InvalidRoundException, \
    InvalidCharactersException, CIEmptyException
from .models import Vendors, VendorContacts, Modules, Rfis, RfiParticipation, CompanyGeneralInfoAnswers, \
    CompanyGeneralInfoQuestion, AssignedVendorsAnalysts, Elements, ParentCategories, Categories, Subcategories, \
    SelfScores, SelfDescriptions, Attachments, ElementsAttachments, AnalystNotes, SmScores, RfiParticipationStatus
from .serializers import VendorsCreateSerializer, VendorToFrontSerializer, VendorsCsvSerializer, ModulesSerializer, \
    VendorsManagementListSerializer, VendorManagementUpdateSerializer, VendorContactSerializer, \
    VendorContactCreateSerializer, RfiRoundSerializer, RfiRoundCloseSerializer, VendorModulesListManagementSerializer, \
    RfiParticipationSerializer, RfiParticipationCsvSerializer, RfiParticipationCsvDownloadSerializer, \
    ContactUpdateSerializer, ElementCommonInfoSerializer, AnalystSerializer, \
    VendorActiveToFrontSerializer, DownloadExcelSerializer, ElementInitializeInfoSerializer


class AdministratorDashboard(APIView):

    def get(self, request, format=None):
        vendors = [vendor.vendor_name for vendor in Vendors.objects.all()]
        return Response(vendors)


class AnalystListView(generics.ListAPIView):

    queryset = AssignedVendorsAnalysts.objects.all()
    serializer_class = AnalystSerializer


class FileUploadView(APIView):
    parser_classes = (MultiPartParser, FormParser)

    def put(self, request, format=None):
        if 'file' not in request.data:
            raise ParseError("Empty content")
        f = request.data['file']
        filename = f.name
        if filename.endswith('.csv'):
            try:
                file = default_storage.save(filename, f)
                r = csv_file_parser(file)
                status = 200
            finally:
                default_storage.delete(file)
        else:
            status = 406
            r = {"general_errors": ["Please upload only CSV files"]}
        return Response(r, status=status)


class ExcelFileUploadView(APIView):
    permission_classes = [permissions.AllowAny]
    parser_classes = (MultiPartParser, FormParser)

    def put(self, request, format=None, **kwargs):
        # context need for avoid full excel file parsing during rfi sheet perform, not implemented if parse Company Info
        context = {'rfiid': kwargs.get('rfiid'), 'vendor': kwargs.get('vendor')}

        if 'file' not in request.data:
            raise ParseError("Empty content")
        f = request.data['file']
        filename = f.name
        if filename.endswith('.xlsx') or filename.endswith('.xls'):
            if re.match(r'^SM_\d{4}Q\d_.*?_\d\d[A-Z]\d_\d\.[a-zA-Z]+$', filename):
                try:
                    file = default_storage.save(filename, f)
                    if filename.endswith('.xls'):
                        import pyexcel
                        _f, _ = filename.split('.')
                        pyexcel.save_book_as(file_name=file, dest_file_name=f'{_f}.xlsx')
                        file = default_storage.url(f'{_f}.xlsx')
                    split_name = self.split_file_name(filename)
                    context.update(split_name)
                    r = get_full_excel_file_response(file, context)
                    status = 200
                except InvalidFormatException as e:
                    r = {"general_errors": [e.__str__()]}
                    status = 406
                except InvalidRoundException as e:
                    r = {"general_errors": [e.__str__()]}
                    status = 406
                except InvalidCharactersException as e:
                    r = {"general_errors": [e.__str__()]}
                    status = 406
                except CIEmptyException as e:
                    r = {"general_errors": [e.__str__()]}
                    status = 406
                except Exception as e:
                    r = {"general_errors": [e.__str__()]}
                    status = 406
                finally:
                    default_storage.delete(file)
                    # TODO delete .xls file too (for now delete only converted file .xlsx)
            else:
                status = 406
                r = {"general_errors": ["File name incorrect. Use this template SM_2020Q1_Vendor_20R1_1.xlsx"]}
        else:
            status = 406
            r = {"general_errors": ["Please upload only xlsx or xls files"]}
        return Response(r, status=status)

    @staticmethod
    def split_file_name(filename):
        """
        Processing a file name and obtaining variables from it
        :param filename:
        :return:
        """
        name = os.path.splitext(filename)[0]
        _, year_quart, vendor_name, round, scoring_round = tuple(name.split('_'))
        split_name = {'f_vendor_name': vendor_name, 'f_round': round, 'f_scoring_round': scoring_round}
        if not scoring_round.isdigit():
            raise ParseError("Error file name. Scoring round incorrect")
        if int(scoring_round) not in range(1, 9):
            raise ParseError("Error file name. Scoring round incorrect")
        return split_name


class CsvToDatabase(APIView):
    """
[
    {
        "vendor_name": "Tefstfdstest43",
        "country": "Belarus",
        "nda": "2019-12-24",
        "modules": [
            {
                "module": "Sourcing"
            },
            {
                "module": "SA"
            }
        ],
        "contacts": [
            {
                "email": "jack15621@gmail.com",
                "contact_name": "Jack Jhonson",
                "primary": true
            },
            {
                "email": "j45ack213@gmail.com",
                "contact_name": "",
                "primary": false
            }
        ]
    },
    {
        "vendor_name": "Tesddt7t2test",
        "country": "Canada",
        "nda": "",
        "modules": [
            {
                "module": ""
            }
        ],
        "contacts": [
            {
                "email": "sand45r2a1@gmail.com",
                "contact_name": "Sandra Bullock",
                "primary": true
            },
            {
                "email": "sa1nd54r13a@gmail.com",
                "contact_name": "Sandra Bullock",
                "primary": false
            }
        ]
    }
]
    """

    serializer_class = VendorsCsvSerializer

    def post(self, request, format=None):
        r_data = request.data
        try:
            # implement transaction  - if exception appear during for loop iteration none data save to DB
            with transaction.atomic():
                for data in r_data:
                    if data['nda'] == '':
                        data['nda'] = None
                    for contact in data['contacts']:
                        if contact['email']:
                            contact['email'] = contact['email'].lower()
                    for module in data['modules']:
                        if module['module']:
                            module['module'] = get_object_or_404(Modules, module_name=module['module']).mid
                        else:
                            data.pop('modules')
                    serializer = VendorsCsvSerializer(data=data)
                    serializer.is_valid(raise_exception=True)
                    serializer.save()

        except ValidationError:
            return Response({"errors": (serializer.errors,)},
                            status=status.HTTP_400_BAD_REQUEST)
        else:
            return Response(serializer.data, status=status.HTTP_200_OK)


class VendorsCreateView(APIView):
    """
    parameters {
            "vendor_name": "Uf2r63",
            "country": "Belarus",
            "nda": "2020-12-12",
            "parent": "",
            "contacts": [{"contact_name": "Mrk", "phone": "2373823", "email": "dtrfd@rgmail.com"},
                         {"contact_name": "Uio", "phone": "3q4567", "email": "ddttcyf@gmail.com" }
                         ]
                }

    responses = {
                "vendor_name": "Uf2r63",
                "country": "Belarus",
                "nda": "2020-12-12",
                "parent": "",
                "contacts": [
                    {
                        "contact_name": "Mrk",
                        "phone": "2373823",
                        "email": "dtrfd@rgmail.com"
                    },
                    {
                        "contact_name": "Uio",
                        "phone": "3q4567",
                        "email": "ddttcyf@gmail.com"
                        }
                    ]
                }
    """

    serializer_class = VendorsCreateSerializer

    def post(self, request, *args, **kwargs):
        data = request.data

        try:
            # implement transaction  - if exception appearpyt during for loop iteration none data save to DB
            with transaction.atomic():
                if data['nda'] == '':
                    data['nda'] = None
                for contact in data['contacts']:
                    if contact['email']:
                        contact['email'] = contact['email'].lower()
                serializer = VendorsCreateSerializer(data=data)
                serializer.is_valid(raise_exception=True)
                serializer.save()
        except ValidationError:
            return Response({"errors": (serializer.errors,)},
                            status=status.HTTP_400_BAD_REQUEST)
        else:
            return Response(request.data, status=status.HTTP_200_OK)


class VendorsToFrontView(generics.ListAPIView):
    """
    Get Vendors list for frontend validation
    """
    queryset = Vendors.objects.all()
    serializer_class = VendorToFrontSerializer


class VendorsActiveToFrontView(generics.ListAPIView):
    """
    Get Vendors list for frontend validation
    """
    queryset = Vendors.objects.filter(active=True)
    serializer_class = VendorActiveToFrontSerializer


class ModulesListView(generics.ListAPIView):
    """
    Get list of all exist modules
    """
    queryset = Modules.objects.all()
    serializer_class = ModulesSerializer


# <--VENDOR PROFILE-->
class VendorProfilePageView(generics.RetrieveAPIView):
    """ Get vendor Profile personal data"""

    serializer_class = VendorContactSerializer
    lookup_field = "vendorid"

    def get_object(self):
        obj = get_object_or_404(VendorContacts, vendor=self.kwargs["vendorid"], primary=True)
        return obj


class VendorManagementListScreen(generics.ListAPIView):
    """n
    Get Vendors Management screen
    """
    serializer_class = VendorsManagementListSerializer
    queryset = Vendors.objects.all()


class VendorProfileUpdateView(generics.RetrieveUpdateAPIView):
    """ Update main vendor info (exclude contact)

    GET:
       {
        "vendorid": 118,
        "vendor_name": "rdty",
        "active": true,
        "country": "Belarus",
        "office": "",
        "abr_date": null,
        "nda": "2020-12-12",
        "parent": null,
        "contacts": [],
        "to_vendor": [
            {
                "pk": 3,
                "active": true,
                "m": "Sourcing",
                "rfi": "20R1",
                "vendor": 118,
                "timestamp": "2020-04-06T12:50:14.762535"
            },
            {
                "pk": 4,
                "active": true,
                "m": "SA",
                "rfi": "20R1",
                "vendor": 118,
                "timestamp": "2020-04-06T13:57:33.988896"
            }
        ],
        "history": [
                {
                    "vendor_name": "rty"
                },
                {
                    "vendor_name": "TTT"
                }
            ],
    "current_round_participate": false
    }



    PUT (PATH):
    Possible send partial data (just one field)
    data = {
            "vendor_name":"Forest G"
           }
    """

    # permission_classes = [permissions.AllowAny, ]
    serializer_class = VendorManagementUpdateSerializer
    lookup_field = 'vendorid'

    def get_queryset(self):
        vendorid = self.kwargs['vendorid']
        return Vendors.objects.filter(vendorid=vendorid)

    def put(self, request, *args, **kwargs):
        return self.partial_update(request, *args, **kwargs)


class VendorContactsCreateView(generics.CreateAPIView):
    """
          {
        "vendor": 1,
        "contact_name": "San",
        "phone": null,
        "email": "tesdвbxxxxdoxxset@gmail.com",
        "primary": false
      }

    """

    serializer_class = VendorContactCreateSerializer

    def create(self, request, *args, **kwargs):
        serializer = self.get_serializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        self.perform_create(serializer)
        headers = self.get_success_headers(serializer.data)
        return Response(serializer.data, status=status.HTTP_200_OK, headers=headers)


class ContactsUpdateView(generics.RetrieveUpdateDestroyAPIView):
    """
    Get or update exist contact at Vendor Management Screen
    Possible send partial data (just one field) (PUT or PATH method)
    data = {
        "contact_name": "Name"
        }

    """

    serializer_class = ContactUpdateSerializer
    lookup_field = 'contact_id'
    queryset = VendorContacts.objects.all()

    # def put(self, request, *args, **kwargs):
    #     return self.partial_update(request, *args, **kwargs)


class VendorProfileModulesListCreate(generics.ListCreateAPIView):
    """View Vendor profile modules activity status and update it
    POST:
        data = {
            "active": false,
            "rfi": "20R1", # allow send blank field (Null)
            "vendor": 15,
            "m": 4
            }
        """

    serializer_class = RfiParticipationSerializer

    # Implement just to rewrite status code from 201(default) to 200
    def create(self, request, *args, **kwargs):
        serializer = self.get_serializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        self.perform_create(serializer)
        headers = self.get_success_headers(serializer.data)
        return Response(serializer.data, status=status.HTTP_200_OK, headers=headers)

    def get_queryset(self, **kwargs):
        round = Rfis.objects.all().order_by('-timestamp').first()
        vendor = get_object_or_404(Vendors, vendorid=self.kwargs['vendorid'])
        queryset = RfiParticipation.objects.filter(vendor=vendor, rfi=round)
        return queryset


# <--RFI-->

class NewRfiRoundCreateView(generics.ListCreateAPIView):

    """New RFI round crete

        data = {
                "issue_datetime": "2020-03-25T10:52:49.677955Z",
                "open_datetime": "2020-03-10T16:06:01+03:00",
                "close_datetime": "2020-03-10T16:06:01+03:00"
               }

    """

    serializer_class = RfiRoundSerializer
    queryset = Rfis.objects.filter(active=True)


class RfiRoundClose(generics.RetrieveUpdateAPIView):
    """
    Close rfi round
    """

    serializer_class = RfiRoundCloseSerializer
    queryset = Rfis.objects.all()
    lookup_field = 'rfiid'

    def put(self, request, *args, **kwargs):
        return self.partial_update(request, *args, **kwargs)


class RfiRoundView(generics.RetrieveAPIView):
    """ Rfi round info
    """

    serializer_class = RfiRoundSerializer
    queryset = Rfis.objects.all()
    lookup_field = 'rfiid'


class RfiRoundListView(generics.ListAPIView):
    """ Rfi round info
    """

    serializer_class = RfiRoundSerializer
    queryset = Rfis.objects.filter(active=True)


class AssociateModulesWithVendorView(generics.ListCreateAPIView):
    """
    RFI: List of vendors with participated modules and modules status change method

    POST request data:
    data =  {
                "active": true,
                "m": 1,
                "rfi": "20R1",
                "vendor": 15
             }

    """
    serializer_class = VendorModulesListManagementSerializer
    queryset = Vendors.objects.all()

    def get_serializer_class(self):
        if self.request.method == 'POST':
            return RfiParticipationSerializer
        return VendorModulesListManagementSerializer

    def get_serializer_context(self):
        """
        Extra context provided to the serializer class.
        """
        context = super().get_serializer_context()
        context.update({"rfiid": self.kwargs['rfiid']})
        return context


# Rfi Csv
class RfiCsvUploadView(APIView):
    """ Upload rfi csv file"""
    parser_classes = (MultiPartParser, FormParser)
    serializer_class = VendorsCsvSerializer

    def put(self, request, format=None):
        if 'file' not in request.data:
            raise ParseError("Empty content")
        f = request.data['file']
        filename = f.name
        if filename.endswith('.csv'):
            try:
                file = default_storage.save(filename, f)
                r = rfi_csv_file_parser(file)
                status = 200
            finally:
                default_storage.delete(file)
        else:
            status = 406
            r = {"general_errors": ["Please upload only CSV files"]}
        return Response(r, status=status)


class AssociateModulesWithVendorCsv(APIView):

    """
    Create or update modules to rfi

    data =
            [
                [
                    {
                        "rfi": "20R1",
                        "vendor": "Actual2",
                        "m": "Strategic Sourcing",
                        "active": "False"
                    },
                    {
                        "rfi": "20R1",
                        "vendor": "Actual2",
                        "m": "Supplier Management",
                        "active": "True"
                    }

                ],
                [
                    {
                        "rfi": "20R1",
                        "vendor": "Test",
                        "m": "Strategic Sourcing",
                        "active": "False"
                    },
                    {
                        "rfi": "20R1",
                        "vendor": "Test",
                        "m": "Supplier Management",
                        "active": "True"
                    }

                ]

            ]
    """
    serializer_class = RfiParticipationCsvSerializer

    def post(self, request, format=None):
        r_data = request.data
        try:
            # implement transaction  - if exception appear during for loop iteration none data save to DB
            with transaction.atomic():
                for data in r_data:
                    for data in data:
                        v = data.get('vendor', None)
                        vendor = get_object_or_404(Vendors, vendor_name=v)
                        data['vendor'] = vendor.vendorid
                        m = data.get('m', None)
                        module = get_object_or_404(Modules, module_name=m)
                        data['m'] = module.mid
                        serializer = RfiParticipationCsvSerializer(data=data)
                        serializer.is_valid(raise_exception=True)
                        serializer.save()
        except ValidationError:
            return Response({"errors": (serializer.errors,)},
                            status=status.HTTP_400_BAD_REQUEST)
        else:
            return Response(request.data, status=status.HTTP_200_OK)


class CsvRfiTemplateDownload(APIView):
    permission_classes = [permissions.AllowAny]

    def get(self, request, format=None, **kwargs):
        # Get list of exist round in db
        rfi = kwargs['rfiid']
        rounds = list(Rfis.objects.filter(active=True).values('rfiid', ))
        round_list = [d.get('rfiid') for d in rounds]
        if rfi not in round_list:
            raise ParseError(detail={"general_errors": ["Round {} is not exist".format(rfi)]}, code=406)

        csv_header_fields = ['Round', 'Vendor', 'Strategic Sourcing', 'Supplier Management', 'Spend Analytics',
                             'Contract Management', 'e-Procurement', 'Invoice-to-Pay', 'AP Automation',
                             'Strategic Procurement Technologies', 'Procure-to-Pay', 'Source-to-Pay']
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = 'attachment; filename="export_rfi.csv"'
        writer = csv.DictWriter(response, fieldnames=csv_header_fields)
        writer.writeheader()
        vendor = Vendors.objects.filter()
        for v in vendor:
            vendor_name = v.vendor_name
            module = RfiParticipation.objects.filter(rfi=rfi, vendor=v)
            module_to_vendor = []
            for m in module:
                serializer = RfiParticipationCsvDownloadSerializer(m)
                module_dict = serializer.data.copy()  # get dict object
                module_to_vendor.append(module_dict)
            res = {i['m']: i['active'] for i in module_to_vendor if i.keys() == {'active', 'm'}}
            writer.writerow({'Round': rfi, 'Vendor': vendor_name,
                             'Strategic Sourcing': res.get('Strategic Sourcing', False),
                             'Supplier Management': res.get('Supplier Management', False),
                             'Spend Analytics': res.get('Spend Analytics', False),
                             'Contract Management': res.get('Contract Management', False),
                             'e-Procurement': res.get('e-Procurement', False),
                             'Invoice-to-Pay': res.get('Invoice-to-Pay', False),
                             'AP Automation': res.get('AP Automation', False),
                             'Strategic Procurement Technologies': res.get('Strategic Procurement Technologies', False),
                             'Procure-to-Pay': res.get('Procure-to-Pay', False),
                             'Source-to-Pay': res.get('Source-to-Pay', False)})
        return response

# EXCEL


class UploadElementFromExcelFile(APIView):
    permission_classes = [permissions.AllowAny]
    serializer_class = ElementCommonInfoSerializer

    def post(self, request, *args, **kwargs):
        context = {'rfiid': kwargs.get('rfiid'), 'vendor': kwargs.get('vendor'), 'analyst': kwargs.get('analyst')}
        data = request.data  # data is list of dict

        for num, _d in enumerate(data):
            if 'Scoring_round_current' in _d:
                current_scoring_round = _d['Scoring_round_current']
                context['current_scoring_round'] = current_scoring_round
                del data[num]
                break

        for num, _d in enumerate(data):
            if 'Company_info' in _d:
                company_information = _d['Company_info']
                context['Company_info'] = company_information
                del data[num]
                break

        # if not kwargs.get('analyst'):  # Check that file send by vendor
        #     # Check that CI answer stored in DB yet (at list one)
        #     ci_db = self.check_ci_exist_in_db(round=kwargs.get('rfiid'),
        #                                       vendor=Vendors.objects.get(vendorid=kwargs.get('vendor')))
        #
        #     # Check company info from excel file
        #     ci_file = self.get_ci_from_excel_file(company_information)
        #     if not ci_db and not ci_file:
        #         r = {"general_errors": ["The company information is blank"]}
        #         return Response(r, status=406)

        for num, _d in enumerate(data):
            if 'Scoring_round_info' in _d:
                pc_status = _d['Scoring_round_info']
                context['pc_status_info'] = pc_status
                del data[num]
                break

        for num, _d in enumerate(data):
            if 'Status_info' in _d:
                status_info = _d['Status_info']
                context['status_info'] = status_info
                del data[num]
                break
        # For CI creation
        round = Rfis.objects.get(rfiid=kwargs.get('rfiid'))
        vendor = Vendors.objects.get(vendorid=kwargs.get('vendor'))

        try:
            # implement transaction  - if exception appear during for loop iteration none data save to DB
            with transaction.atomic():
                # CI creations
                for ci in company_information:
                    ciq, _ = CompanyGeneralInfoQuestion.objects.get_or_create(question=ci.get('question'), rfi=round)
                    cia, _ = CompanyGeneralInfoAnswers.objects.update_or_create(vendor=vendor, question=ciq,
                                                                                scoring=current_scoring_round,
                                                                                defaults={'answer': ci.get('answer')})
                for pc_data in data:
                    parent_category = pc_data.get('Parent Category')
                    parent_category = parent_category.rstrip()
                    category_data = pc_data.get('Category')
                    for data in category_data:
                        for category, values in data.items():  # Get category name
                            for subcats in values:
                                for subcat, element_list in subcats.items():  # Get subcategory name
                                    for num, element in enumerate(element_list, 1):  # Get element info
                                        element_name = element.get('Element name')
                                        e_order = num
                                        category = category
                                        pc = parent_category
                                        s = subcat
                                        description = element.get('Description')
                                        scoring_scale = element.get('Scoring Scale')
                                        self_score = element.get('Self-Score')
                                        self_description = element.get('Self-Description')
                                        sm_score = element.get('SM score')
                                        analyst_notes = element.get('Analyst notes')
                                        attachment = element.get('Attachments/Supporting Docs and Location/Link')
                                        data = {'element_name': element_name, 'e_order': e_order,
                                                'description': description, 'scoring_scale': scoring_scale,
                                                'self_score': self_score, 'self_description': self_description,
                                                'sm_score': sm_score, 'analyst_notes': analyst_notes,
                                                'attachment': attachment, 's': s, 'category': category, 'pc': pc
                                                }
                                        serializer = ElementCommonInfoSerializer(data=data, context=context)
                                        serializer.is_valid(raise_exception=True)
                                        serializer.save()
        except ValidationError:
            return Response({"errors": (serializer.errors,)},
                            status=status.HTTP_400_BAD_REQUEST)
        else:
            return Response(request.data, status=status.HTTP_200_OK)

    # @staticmethod
    # def check_ci_exist_in_db(round, vendor):
    #     """Check that CI answer stored in DB yet (at list one)"""
    #     round = round
    #     exist_company_question = CompanyGeneralInfoQuestion.objects.filter(rfi=round)
    #     ci_exist = False
    #
    #     if exist_company_question:
    #         for q in exist_company_question:
    #             if q.answer_to_question.filter(vendor=vendor):  # check if answer is exist and not None
    #                 _a = (q.answer_to_question.filter(vendor=vendor).first())
    #                 if _a.answer:
    #                     ci_exist = True
    #     return ci_exist
    #
    # @staticmethod
    # def get_ci_from_excel_file(company_information):
    #     """"Check company info from excel file"""
    #     # company_information = next(iter(data))  # company information as a dict
    #     # information = company_information.get('Company_info')
    #     ci_exist = False
    #     for i in company_information:
    #         if i.get('answer'):
    #             ci_exist = True
    #     return ci_exist

    # @staticmethod
    # def not_all_element_is_null(data):
    #     """
    #     Check that at list one element pair (self_score/self_description; sm_score/analyst_notes) are not empty.
    #     That means we can set rfi_part_status to PC as positive digit(1 for first scoring round etc.)
    #     :param data:
    #     :return:
    #     """
    #     from_vendor = tuple(data.get('self_score'), data.get('self_description'))
    #     from_analytic = tuple(data.get('sm_score', data.get('analyst_notes')))
    #     if not all(from_vendor) and not all(from_analytic):
    #         return True
    #     else:
    #         return False


class InfoToDownloadRfiExcelFile(generics.ListAPIView):
    """
    GET:
       return list of active vendor with last scoring round
    """
    permission_classes = [permissions.AllowAny]
    serializer_class = DownloadExcelSerializer
    model = serializer_class.Meta.model

    def get_queryset(self):
        queryset = self.model.objects.filter(active=True)
        return queryset

    def get_serializer_context(self):
        """
        Extra context provided to the serializer class.
        """
        context = super().get_serializer_context()
        context.update({"rfiid": self.kwargs['rfiid']})
        return context

class DownloadRfiExcelFile(APIView):
    permission_classes = [permissions.AllowAny]
    # """
    # Prepare empty template
    # """
    # def get(self, request, format=None, **kwargs):
    #     file = default_storage.url('test.xlsx')
    #     wb = load_workbook(filename=file)
    #
    #     response = HttpResponse(content_type='application/vnd.ms-excel')
    #     response['Content-Disposition'] = 'attachment; filename="test.xlsx"'
    #     sheet = wb["RFI"]
    #     wb.remove_sheet(sheet)
    #     wb.create_sheet('RFI')
    #     # # Setting initial values for deletion
    #     # starting_row = 4
    #     # starting_col = self.col2num('U')
    #     # last_col = self.col2num('Y')
    #     # # Deleting rows and columns
    #     # sheet.delete_rows(starting_row, sheet.max_row - starting_row)
    #     # sheet.delete_cols(starting_col, last_col - starting_col + 1)
    #     # merged_cell_coord = []
    #     # for range_ in sheet.merged_cell_ranges:
    #     #     # get current coordinate from all merget cell and set it as a string
    #     #     merged_cell_coord.append(range_.__str__())
    #     # for i in (merged_cell_coord):
    #     #     print(i)
    #     #     print(type(i))
    #     #     sheet.unmerge_cells(i)
    #
    #     wb.save(response)
    #     return response
    #
    # @staticmethod
    # def col2num(col):
    #     # Utility function to convert culomn letters to numbers
    #     num = 0
    #     for c in col:
    #         if c in string.ascii_letters:
    #             num = num * 26 + (ord(c.upper()) - ord('A')) + 1
    #     return num

    def post(self, request, format=None, **kwargs):

        """
        :param request:
        {
                "rfiid": "20R1",
                "vendorid": 122,
                "scoring_status": 1
             }
        :param format:
        :param kwargs:
        :return:
        """

        # TODO check that element and self_score, self_description ...  are not empty (exist in DB) ???
        # list{	"rfiid": "20R1", 'vendorid': 122, 'vendor_name': 'Actual2', 'scoring_status': 3}
        for request in request.data:
            r = request
            if r.get('vendorid') is None:
                continue

            rfi = Rfis.objects.get(rfiid=r.get('rfiid'))
            vendor = Vendors.objects.get(vendorid=r.get('vendorid'))
            current_scoring_round = r.get('scoring_status')

            if current_scoring_round is None:
                status = 406
                r = {"general_errors": ["The vendor doesn't have active module in round."]}
                return Response(r, status=status)

            if not RfiParticipation.objects.filter(vendor=vendor, rfi=rfi, active=True):
                status = 406
                r = {"general_errors": ["The vendor doesn't have active module in current round."]}
                return Response(r, status=status)

            # Get vendor active module and calculate participate PC
            participate_module = RfiParticipation.objects.filter(vendor=vendor, rfi=rfi, active=True)
            participate_module_list = [element.m.module_name for element in participate_module]
            unique_pc = list(get_excel_file_current_pc_for_parsing(pml=participate_module_list))  # Get unique PC for future processing
            if not unique_pc:
                status = 406
                r = {"general_errors": ["The vendor doesn't have parent category in current round."]}
                return Response(r, status=status)
            file = default_storage.url('blank_template.xlsx')
            wb = load_workbook(filename=file)
            ws = wb["RFI"]

            # Add column size
            column_dimensions = ws.column_dimensions['E']
            column_dimensions.width = 40
            column_dimensions = ws.column_dimensions['F']
            column_dimensions.width = 50
            column_dimensions = ws.column_dimensions['G']
            column_dimensions.width = 70
            column_dimensions = ws.column_dimensions['Q']
            column_dimensions.width = 60
            column_dimensions = ws.column_dimensions['R']
            column_dimensions.width = 40
            column_dimensions = ws.column_dimensions['T']
            column_dimensions.width = 60
            column_dimensions = ws.column_dimensions['V']
            column_dimensions.width = 60
            column_dimensions = ws.column_dimensions['W']
            column_dimensions.width = 40
            column_dimensions = ws.column_dimensions['Y']
            column_dimensions.width = 60

            # Style variables
            thin_border = Border(left=Side(style='thin'),
                                 right=Side(style='thin'),
                                 top=Side(style='thin'),
                                 bottom=Side(style='thin'))
            cell_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            element_alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)

            #Create header
            ws['E2'] = "Element name"
            ws['E2'].alignment = cell_alignment
            ws['E2'].fill = PatternFill(start_color="61A144", fill_type="solid")
            ws['E2'].border = thin_border
            ws['F2'] = "Description"
            ws['F2'].alignment = cell_alignment
            ws['F2'].fill = PatternFill(start_color="61A144", fill_type="solid")
            ws['F2'].border = thin_border
            ws['G2'] = 'Scoring Scale'
            ws['G2'].alignment = cell_alignment
            ws['G2'].fill = PatternFill(start_color="61A144", fill_type="solid")
            ws['G2'].border = thin_border
            ws['P2'] = 'Self-Score'
            ws['P2'].alignment = cell_alignment
            ws['P2'].fill = PatternFill(start_color="B2F4FF", fill_type="solid")
            ws['P2'].border = thin_border
            ws['Q2'] = 'Self-Description'
            ws['Q2'].alignment = cell_alignment
            ws['Q2'].fill = PatternFill(start_color="B2F4FF", fill_type="solid")
            ws['Q2'].border = thin_border
            ws['R2'] = 'Attachments/Supporting Docs and Location/Link'
            ws['R2'].alignment = cell_alignment
            ws['R2'].fill = PatternFill(start_color="B2F4FF", fill_type="solid")
            ws['R2'].border = thin_border
            ws['S2'] = 'SM score'
            ws['S2'].alignment = cell_alignment
            ws['S2'].fill = PatternFill(start_color="EDBC00", fill_type="solid")
            ws['S2'].border = thin_border
            ws['T2'] = 'Analyst notes'
            ws['T2'].alignment = cell_alignment
            ws['T2'].fill = PatternFill(start_color="EDBC00", fill_type="solid")
            ws['T2'].border = thin_border
            ws['AE2'] = 'Current Self-Score'
            ws['AE2'].alignment = cell_alignment
            ws['AE2'].fill = PatternFill(start_color="FFE5EA", fill_type="solid")
            ws['AE2'].border = thin_border
            ws['AF2'] = 'Current score'
            ws['AF2'].alignment = cell_alignment
            ws['AF2'].fill = PatternFill(start_color="61A144", fill_type="solid")
            ws['AF2'].border = thin_border

            # 2-d scoring_round

            ws['U2'] = 'Self-Score (2)'
            ws['U2'].alignment = cell_alignment
            ws['U2'].fill = PatternFill(start_color="B2F4FF", fill_type="solid")
            ws['U2'].border = thin_border
            ws['V2'] = 'Reasoning'
            ws['V2'].alignment = cell_alignment
            ws['V2'].fill = PatternFill(start_color="B2F4FF", fill_type="solid")
            ws['V2'].border = thin_border
            ws['W2'] = 'Attachments/Supporting Docs and Location/Link'
            ws['W2'].alignment = cell_alignment
            ws['W2'].fill = PatternFill(start_color="B2F4FF", fill_type="solid")
            ws['W2'].border = thin_border
            ws['X2'] = 'SM score (2)'
            ws['X2'].alignment = cell_alignment
            ws['X2'].fill = PatternFill(start_color="EDBC00", fill_type="solid")
            ws['X2'].border = thin_border
            ws['Y2'] = 'Analyst notes (2)'
            ws['Y2'].alignment = cell_alignment
            ws['Y2'].fill = PatternFill(start_color="EDBC00", fill_type="solid")
            ws['Y2'].border = thin_border

            # 3-d scoring_round

            ws['Z2'] = 'Self-Score (3)'
            ws['Z2'].alignment = cell_alignment
            ws['Z2'].fill = PatternFill(start_color="B2F4FF", fill_type="solid")
            ws['Z2'].border = thin_border
            ws['AA2'] = 'Reasoning'
            ws['AA2'].alignment = cell_alignment
            ws['AA2'].fill = PatternFill(start_color="B2F4FF", fill_type="solid")
            ws['AA2'].border = thin_border
            ws['AB2'] = 'Attachments/Supporting Docs and Location/Link'
            ws['AB2'].alignment = cell_alignment
            ws['AB2'].fill = PatternFill(start_color="B2F4FF", fill_type="solid")
            ws['AB2'].border = thin_border
            ws['AC2'] = 'SM score (3)'
            ws['AC2'].alignment = cell_alignment
            ws['AC2'].fill = PatternFill(start_color="EDBC00", fill_type="solid")
            ws['AC2'].border = thin_border
            ws['AD2'] = 'Analyst notes (3)'
            ws['AD2'].alignment = cell_alignment
            ws['AD2'].fill = PatternFill(start_color="EDBC00", fill_type="solid")
            ws['AD2'].border = thin_border

            ws.row_dimensions[2].height = 60

            for rows in ws.iter_rows(min_row=3, max_row=3, min_col=5, max_col=7):
                for cell in rows:
                    cell.fill = PatternFill(start_color="790099", fill_type="solid")

            # Hidden column
            for col in ['B', 'C', 'D', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O', 'U', 'V', 'W', 'X', 'Y', 'Z', 'AA', 'AB',
                        'AC', 'AD']:
                ws.column_dimensions[col].hidden = True

            row_num = 4

            if current_scoring_round == 1:
                if self.catch_zero_round(vendor, rfi):
                    self.initialize_template_create(ws, unique_pc, row_num, cell_alignment, thin_border, element_alignment)

            # Start excel RFI sheet create logic
            if not self.catch_zero_round(vendor, rfi):
                for pc in unique_pc:
                    ws[f'E{row_num}'] = pc
                    ws[f'E{row_num}'].alignment = cell_alignment
                    ws[f'E{row_num}'].border = thin_border
                    ws.merge_cells(f'E{row_num}:G{row_num}')
                    ws.row_dimensions[row_num].height = 40
                    for rows in ws.iter_rows(min_row=row_num, max_row=row_num, min_col=5, max_col=7):
                        for cell in rows:
                            cell.fill = PatternFill(start_color="EDBC00", fill_type="solid")
                    row_num += 1
                    p_c = ParentCategories.objects.get(parent_category_name=pc)
                    categories = Categories.objects.filter(pc=p_c)
                    for category in categories:
                        # ws.merge_cells(f'E{row_num}:G{row_num}') # get  "'MergedCell' object has no attribute 'column_letter'"
                        ws[f'E{row_num}'] = category.category_name
                        ws[f'E{row_num}'].alignment = cell_alignment
                        ws[f'E{row_num}'].border = thin_border
                        for rows in ws.iter_rows(min_row=row_num, max_row=row_num, min_col=5, max_col=7):
                            for cell in rows:
                                cell.fill = PatternFill(start_color="61A144", fill_type="solid")
                        ws.row_dimensions[row_num].height = 25
                        row_num += 1
                        subcats = Subcategories.objects.filter(c=category).order_by('timestamp')
                        for subcat in subcats:
                            if not subcat.subcategory_name == 'General':
                                ws[f'E{row_num}'] = subcat.subcategory_name
                                ws[f'E{row_num}'].fill = PatternFill(start_color="E5FBFF", fill_type="solid")
                                ws[f'E{row_num}'].border = thin_border
                                row_num += 1
                            elements = Elements.objects.filter(s=subcat).order_by('e_order')
                            for e in elements:
                                column_to_scoring_round = {
                                                           '1': ['E', 'F', 'G', 'P', 'Q', 'R', 'S', 'T'],
                                                           '2': ['E', 'F', 'G', 'U', 'V', 'W', 'X', 'Y'],
                                                           '3': ['E', 'F', 'G', 'Z', 'AA', 'AB', 'AC', 'AD']
                                                            }

                                ws.row_dimensions[row_num].height = 150

                                # Element common info create
                                element_name = e.element_name
                                description = e.description
                                s_scale = e.scoring_scale

                                ws[f'{column_to_scoring_round.get(str(current_scoring_round))[0]}{row_num}'] = element_name
                                ws[f'{column_to_scoring_round.get(str(current_scoring_round))[0]}{row_num}'].alignment = element_alignment
                                ws[f'{column_to_scoring_round.get(str(current_scoring_round))[1]}{row_num}'] = description
                                ws[f'{column_to_scoring_round.get(str(current_scoring_round))[1]}{row_num}'].alignment = element_alignment
                                ws[f'{column_to_scoring_round.get(str(current_scoring_round))[2]}{row_num}'] = s_scale
                                ws[f'{column_to_scoring_round.get(str(current_scoring_round))[2]}{row_num}'].alignment = element_alignment

                                for csr in range(1, current_scoring_round + 1):

                                    self_score = SelfScores.objects.filter(e=e, vendor=vendor, rfi=rfi, vendor_response=csr).first()
                                    if self_score:
                                        self_score = self_score.self_score
                                    else:
                                        self_score = None

                                    self_description = SelfDescriptions.objects.filter(e=e, vendor=vendor, rfi=rfi, vendor_response=csr).first()
                                    if self_description:
                                        self_description = self_description.self_description
                                    else:
                                        self_description = None

                                    sm_score = SmScores.objects.filter(e=e, vendor=vendor, rfi=rfi, analyst_response=csr).first()
                                    if sm_score:
                                        sm_score = sm_score.sm_score
                                    else:
                                        sm_score = None

                                    analyst_notes = AnalystNotes.objects.filter(e=e, vendor=vendor, rfi=rfi, analyst_response=csr).first()
                                    if analyst_notes:
                                        analyst_notes = analyst_notes.analyst_notes
                                    else:
                                        analyst_notes = None

                                    attachment = ElementsAttachments.objects.filter(e=e, vendor=vendor, rfi=rfi, vendor_response=csr).first()
                                    if attachment:
                                        attachment = attachment.attachment_info
                                    else:
                                        attachment = None
                                    ws[f'{column_to_scoring_round.get(str(csr))[3]}{row_num}'] = self_score
                                    ws[f'{column_to_scoring_round.get(str(csr))[3]}{row_num}'].alignment = element_alignment
                                    ws[f'{column_to_scoring_round.get(str(csr))[4]}{row_num}'] = self_description
                                    ws[f'{column_to_scoring_round.get(str(csr))[4]}{row_num}'].alignment = element_alignment
                                    ws[f'{column_to_scoring_round.get(str(csr))[5]}{row_num}'] = attachment
                                    ws[f'{column_to_scoring_round.get(str(csr))[5]}{row_num}'].alignment = element_alignment
                                    ws[f'{column_to_scoring_round.get(str(csr))[6]}{row_num}'] = sm_score
                                    ws[f'{column_to_scoring_round.get(str(csr))[6]}{row_num}'].alignment = element_alignment
                                    ws[f'{column_to_scoring_round.get(str(csr))[7]}{row_num}'] = analyst_notes
                                    ws[f'{column_to_scoring_round.get(str(csr))[7]}{row_num}'].alignment = element_alignment
                                row_num += 1
                            row_num += 2  # two empty row after subcategory block

                # CI filling
                ws_ci = wb["Company Information"]
                start_cell_row_number = 5
                ciq_queriset = CompanyGeneralInfoQuestion.objects.filter(rfi=rfi)
                for ciq in ciq_queriset:
                    cia = CompanyGeneralInfoAnswers.objects.filter(vendor=vendor, question=ciq).first()
                    ws_ci[f'C{start_cell_row_number}'] = cia.answer
                    start_cell_row_number += 1

            # Unmerge column
            if current_scoring_round == 2:
                for col in ['U', 'V', 'W', 'X', 'Y']:
                    ws.column_dimensions[col].hidden = False
            if current_scoring_round == 3:
                for col in ['U', 'V', 'W', 'X', 'Y', 'Z', 'AA', 'AB', 'AC', 'AD']:
                    ws.column_dimensions[col].hidden = False

            # Generate file name
            new_file_name = self.generate_file_name(rfi, vendor, current_scoring_round)

            # wb.save(filename=new_file_name)
            path_to_temp_folder = os.path.dirname(BASE_DIR)
            if not os.path.exists(f'{path_to_temp_folder}/rfi'):
                pathlib.Path(f'{path_to_temp_folder}/rfi').mkdir(parents=True, exist_ok=True)
            wb.save(f'{path_to_temp_folder}/rfi/{new_file_name}')

        archive = self.generate_zip_name(rfi)
        to_rar = f'{path_to_temp_folder}/rfi'
        to_download = f'{path_to_temp_folder}/{archive}'
        # patoolib.create_archive(to_download, (to_rar,))
        _, to_rar = (os.path.split(to_rar))
        os.system("rar a {} {}".format(archive, to_rar))
        if os.path.exists(to_download):
            try:
                with open(to_download, 'rb') as fh:
                    response = HttpResponse(fh.read(),
                                            content_type="content_type='application/vnd.rar'")
                    response['Content-Disposition'] = 'attachment; filename= "{}"'.format(archive)
                    return response
            finally:
                shutil.rmtree(to_rar, ignore_errors=True)
                default_storage.delete(to_download)

        else:
            raise ParseError

    @staticmethod
    def catch_zero_round(vendor, rfi):
        max_score = RfiParticipationStatus.objects.filter(vendor=vendor, rfi=rfi).aggregate(Max('last_vendor_response'),
                                                                                            Max('last_analyst_response'))
        if max_score.get('last_vendor_response__max') is None and max_score.get('last_analyst_response__max') is None:
            return True
        if max_score.get('last_vendor_response__max') == 0 and max_score.get('last_analyst_response__max') == 0:
            return True
        else:
            return False

    @staticmethod
    def generate_file_name(rfi, vendor, current_scoring_round):
        RFI_Round = rfi.rfiid
        v_name = vendor.vendor_name
        vendor_name = ''.join(v_name.split())
        version = current_scoring_round

        year_ = str(date.today().year)
        month_ = str(date.today().month)
        q_ = {'Q1': ['1', '2', '3'], 'Q2': ['4', '5', '6'], 'Q3': ['7', '8', '9'], 'Q4': ['10', '11', '12']}
        q = 0
        for k, v in q_.items():
            if month_ in v:
                q = k

        return f'SM_{year_}{q}_{vendor_name}_{RFI_Round}_{version}.xlsx'

    @staticmethod
    def generate_zip_name(rfi):
        RFI_Round = rfi.rfiid

        year_ = str(date.today().year)
        month_ = str(date.today().month)
        q_ = {'Q1': ['1', '2', '3'], 'Q2': ['4', '5', '6'], 'Q3': ['7', '8', '9'], 'Q4': ['10', '11', '12']}
        q = 0
        for k, v in q_.items():
            if month_ in v:
                q = k

        return f'SM_{year_}{q}_{RFI_Round}.rar'


    @staticmethod
    def initialize_template_create(ws, unique_pc, row_num, cell_alignment, thin_border, element_alignment):
        # filling empty template
        for pc in unique_pc:
            ws[f'E{row_num}'] = pc
            ws[f'E{row_num}'].alignment = cell_alignment
            ws[f'E{row_num}'].border = thin_border
            ws.merge_cells(f'E{row_num}:G{row_num}')
            ws.row_dimensions[row_num].height = 40
            for rows in ws.iter_rows(min_row=row_num, max_row=row_num, min_col=5, max_col=7):
                for cell in rows:
                    cell.fill = PatternFill(start_color="EDBC00", fill_type="solid")
            row_num += 1
            p_c = ParentCategories.objects.get(parent_category_name=pc)
            categories = Categories.objects.filter(pc=p_c).order_by('timestamp')
            for category in categories:
                # ws.merge_cells(f'E{row_num}:G{row_num}') # get  "'MergedCell' object has no attribute 'column_letter'"
                ws[f'E{row_num}'] = category.category_name
                ws[f'E{row_num}'].alignment = cell_alignment
                ws[f'E{row_num}'].border = thin_border
                for rows in ws.iter_rows(min_row=row_num, max_row=row_num, min_col=5, max_col=7):
                    for cell in rows:
                        cell.fill = PatternFill(start_color="61A144", fill_type="solid")
                ws.row_dimensions[row_num].height = 25
                row_num += 1
                subcats = Subcategories.objects.filter(c=category).order_by('timestamp')
                for subcat in subcats:
                    if not subcat.subcategory_name == 'General':
                        ws[f'E{row_num}'] = subcat.subcategory_name
                        ws[f'E{row_num}'].fill = PatternFill(start_color="E5FBFF", fill_type="solid")
                        ws[f'E{row_num}'].border = thin_border
                        row_num += 1
                    elements = Elements.objects.filter(s=subcat).order_by('e_order')
                    for e in elements:
                        ws.row_dimensions[row_num].height = 150
                        element_name = e.element_name
                        description = e.description
                        s_scale = e.scoring_scale

                        self_score = None
                        self_description = None
                        sm_score = None
                        analyst_notes = None
                        attachment = None

                        ws[f'E{row_num}'] = element_name
                        ws[f'E{row_num}'].alignment = element_alignment
                        ws[f'F{row_num}'] = description
                        ws[f'F{row_num}'].alignment = element_alignment
                        ws[f'G{row_num}'] = s_scale
                        ws[f'G{row_num}'].alignment = element_alignment
                        ws[f'P{row_num}'] = self_score
                        ws[f'P{row_num}'].alignment = element_alignment
                        ws[f'Q{row_num}'] = self_description
                        ws[f'Q{row_num}'].alignment = element_alignment
                        ws[f'R{row_num}'] = attachment
                        ws[f'R{row_num}'].alignment = element_alignment
                        ws[f'S{row_num}'] = sm_score
                        ws[f'S{row_num}'].alignment = element_alignment
                        ws[f'T{row_num}'] = analyst_notes
                        ws[f'T{row_num}'].alignment = element_alignment
                        row_num += 1
                    row_num += 2  # two empty row after subcategory block
            row_num += 2  # two empty row after PC bloc



# Initialization of the zero template creation (only description of elements) for the first vendor upload

class ElementInitializeFromExcelFile(APIView):

    """
    NOT FOR FE USING !!!

    Initialization of the zero template creation (only description of elements) for the first vendor upload
    """

    permission_classes = [permissions.AllowAny]
    serializer_class = ElementInitializeInfoSerializer

    def post(self, request, *args, **kwargs):
        data = request.data  # data is list of dict

        for num, _d in enumerate(data):
            if 'Scoring_round_current' in _d:
                del data[num]
                break

        for num, _d in enumerate(data):
            if 'Company_info' in _d:
                del data[num]
                break

        for num, _d in enumerate(data):
            if 'Scoring_round_info' in _d:
                del data[num]
                break

        for num, _d in enumerate(data):
            if 'Status_info' in _d:
                del data[num]
                break

        try:
            # implement transaction  - if exception appear during for loop iteration none data save to DB
            with transaction.atomic():

                for pc_data in data:
                    parent_category = pc_data.get('Parent Category')
                    parent_category = parent_category.rstrip()
                    category_data = pc_data.get('Category')
                    for data in category_data:
                        for category, values in data.items():  # Get category name
                            for subcats in values:
                                for subcat, element_list in subcats.items():  # Get subcategory name
                                    for num, element in enumerate(element_list, 1):  # Get element info
                                        element_name = element.get('Element Name')
                                        e_order = num
                                        category = category
                                        pc = parent_category
                                        s = subcat
                                        description = element.get('Description')
                                        scoring_scale = element.get('Scoring Scale')
                                        data = {'element_name': element_name, 'e_order': e_order,
                                                'description': description, 'scoring_scale': scoring_scale,
                                                's': s, 'category': category, 'pc': pc
                                                }
                                        serializer = ElementInitializeInfoSerializer(data=data)
                                        serializer.is_valid(raise_exception=True)
                                        serializer.save()

        except ValidationError:
            return Response({"errors": (serializer.errors,)},
                            status=status.HTTP_400_BAD_REQUEST)
        else:
            return Response(request.data, status=status.HTTP_200_OK)
